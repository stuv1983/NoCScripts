#!/usr/bin/env python3
"""Simple Tkinter tool for comparing an older missing-patches report with a current patch report.

The original report is not changed. A new Excel workbook is created with three extra columns:
Current Patch Status, Current Install Date, and Check Result.

Dependency:
    py -m pip install openpyxl
"""

from __future__ import annotations

import csv
import html
import os
import re
import sys
import traceback
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    root = tk.Tk()
    root.withdraw()
    messagebox.showerror(
        "Missing dependency",
        "This tool requires openpyxl.\n\nInstall it with:\npy -m pip install openpyxl",
    )
    raise SystemExit(1)


REQUIRED_OLD = {"client", "device", "patch", "status"}
REQUIRED_CURRENT = {"client", "device", "patch", "status"}

HEADER_ALIASES = {
    "client name": "client",
    "customer": "client",
    "customer name": "client",
    "site name": "site",
    "device name": "device",
    "computer": "device",
    "computer name": "device",
    "hostname": "device",
    "patch name": "patch",
    "update": "patch",
    "update name": "patch",
    "patch status": "status",
    "discovered install date": "discovered / install date",
    "discovered date install date": "discovered / install date",
    "install date": "discovered / install date",
}

STATUS_PRIORITY = {
    "installed": 100,
    "reboot required": 90,
    "installing": 80,
    "pending": 70,
    "missing": 60,
    "failed": 50,
    "ignored": 40,
}

RESOLVED_STATUSES = {"installed", "reboot required"}
OUTSTANDING_STATUSES = {"installing", "pending", "missing", "failed", "ignored"}


class ReportError(Exception):
    """Raised when a selected report cannot be understood."""


def normalise_text(value: Any) -> str:
    text = html.unescape("" if value is None else str(value))
    return re.sub(r"\s+", " ", text.strip()).casefold()


def normalise_header(value: Any) -> str:
    text = normalise_text(value)
    text = re.sub(r"[^a-z0-9]+", " ", text).strip()
    return HEADER_ALIASES.get(text, text)


def patch_key(row: dict[str, Any]) -> tuple[str, str, str]:
    # Site is deliberately excluded: devices sometimes move sites or site spelling changes.
    return (
        normalise_text(row.get("client")),
        normalise_text(row.get("device")),
        normalise_text(row.get("patch")),
    )


def _clean_row(headers: list[str], values: Iterable[Any]) -> dict[str, Any]:
    row: dict[str, Any] = {}
    for header, value in zip(headers, values):
        if header:
            row[header] = value
    return row


def read_csv_report(path: Path) -> tuple[list[str], list[str], list[dict[str, Any]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        sample = handle.read(8192)
        handle.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel

        reader = csv.reader(handle, dialect)
        raw_rows = list(reader)

    if not raw_rows:
        raise ReportError(f"{path.name} is empty.")

    header_index = next((i for i, row in enumerate(raw_rows) if any(str(v).strip() for v in row)), None)
    if header_index is None:
        raise ReportError(f"{path.name} does not contain a header row.")

    display_headers = [str(v).strip() for v in raw_rows[header_index]]
    canonical_headers = [normalise_header(v) for v in display_headers]
    rows = [
        _clean_row(canonical_headers, row)
        for row in raw_rows[header_index + 1 :]
        if any(str(v).strip() for v in row)
    ]
    return display_headers, canonical_headers, rows


def read_excel_report(path: Path) -> tuple[list[str], list[str], list[dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        raw_rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    if not raw_rows:
        raise ReportError(f"{path.name} is empty.")

    header_index = next((i for i, row in enumerate(raw_rows) if any(v not in (None, "") for v in row)), None)
    if header_index is None:
        raise ReportError(f"{path.name} does not contain a header row.")

    display_headers = ["" if v is None else str(v).strip() for v in raw_rows[header_index]]
    canonical_headers = [normalise_header(v) for v in display_headers]
    rows = [
        _clean_row(canonical_headers, row)
        for row in raw_rows[header_index + 1 :]
        if any(v not in (None, "") for v in row)
    ]
    return display_headers, canonical_headers, rows


def read_report(path_text: str) -> tuple[list[str], list[str], list[dict[str, Any]]]:
    path = Path(path_text)
    if not path.is_file():
        raise ReportError(f"File not found: {path}")

    suffix = path.suffix.casefold()
    if suffix == ".csv":
        return read_csv_report(path)
    if suffix in {".xlsx", ".xlsm"}:
        return read_excel_report(path)
    raise ReportError(f"Unsupported file type: {path.suffix}. Use CSV or XLSX.")


def validate_headers(path: str, canonical_headers: list[str], required: set[str]) -> None:
    available = {header for header in canonical_headers if header}
    missing = sorted(required - available)
    if missing:
        raise ReportError(
            f"{Path(path).name} is missing required column(s): {', '.join(missing)}.\n"
            f"Found: {', '.join(h for h in canonical_headers if h)}"
        )


def choose_current_record(records: list[dict[str, Any]]) -> dict[str, Any]:
    return max(
        records,
        key=lambda row: STATUS_PRIORITY.get(normalise_text(row.get("status")), 0),
    )


def classify(current_status: str | None) -> str:
    status = normalise_text(current_status)
    if status == "installed":
        return "NOW INSTALLED"
    if status == "reboot required":
        return "INSTALLED - REBOOT REQUIRED"
    if status in OUTSTANDING_STATUSES:
        return "STILL OUTSTANDING"
    if status:
        return "CURRENT STATUS: " + str(current_status).upper()
    return "NOT FOUND IN CURRENT REPORT"


def compare_reports(old_rows: list[dict[str, Any]], current_rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], Counter]:
    current_index: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in current_rows:
        current_index[patch_key(row)].append(row)

    compared: list[dict[str, Any]] = []
    counts: Counter = Counter()

    for old_row in old_rows:
        matches = current_index.get(patch_key(old_row), [])
        current = choose_current_record(matches) if matches else None
        current_status = current.get("status") if current else ""
        result = classify(current_status)
        counts[result] += 1

        compared.append(
            {
                "old": old_row,
                "current_status": current_status or "Not found",
                "current_date": (current or {}).get("discovered / install date", ""),
                "result": result,
            }
        )

    return compared, counts


def write_output(
    output_path: str,
    display_headers: list[str],
    canonical_headers: list[str],
    compared: list[dict[str, Any]],
    old_path: str,
    current_path: str,
    counts: Counter,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Checked Patches"

    output_headers = display_headers + ["Current Patch Status", "Current Install Date", "Check Result"]
    sheet.append(output_headers)

    fills = {
        "NOW INSTALLED": PatternFill("solid", fgColor="C6EFCE"),
        "INSTALLED - REBOOT REQUIRED": PatternFill("solid", fgColor="DDEBF7"),
        "STILL OUTSTANDING": PatternFill("solid", fgColor="FFC7CE"),
        "NOT FOUND IN CURRENT REPORT": PatternFill("solid", fgColor="FFEB9C"),
    }

    for item in compared:
        old = item["old"]
        original_values = [old.get(canonical, "") for canonical in canonical_headers]
        sheet.append(original_values + [item["current_status"], item["current_date"], item["result"]])
        row_number = sheet.max_row
        fill = fills.get(item["result"])
        if fill:
            for cell in sheet[row_number]:
                cell.fill = fill

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 12), 55)

    summary = workbook.create_sheet("Summary", 0)
    summary.append(["Patch Report Comparison", ""])
    summary.append(["Older report", old_path])
    summary.append(["Current report", current_path])
    summary.append(["Total old missing rows", len(compared)])
    summary.append(["Now installed", counts.get("NOW INSTALLED", 0)])
    summary.append(["Installed - reboot required", counts.get("INSTALLED - REBOOT REQUIRED", 0)])
    summary.append(["Still outstanding", counts.get("STILL OUTSTANDING", 0)])
    summary.append(["Not found in current report", counts.get("NOT FOUND IN CURRENT REPORT", 0)])

    summary["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    summary["B1"].font = Font(bold=True, size=14, color="FFFFFF")
    summary["A1"].fill = header_fill
    summary["B1"].fill = header_fill
    summary.column_dimensions["A"].width = 34
    summary.column_dimensions["B"].width = 90

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)


def process_reports(old_path: str, current_path: str, output_path: str) -> Counter:
    old_display, old_canonical, old_rows = read_report(old_path)
    _, current_canonical, current_rows = read_report(current_path)

    validate_headers(old_path, old_canonical, REQUIRED_OLD)
    validate_headers(current_path, current_canonical, REQUIRED_CURRENT)

    compared, counts = compare_reports(old_rows, current_rows)
    write_output(
        output_path,
        old_display,
        old_canonical,
        compared,
        old_path,
        current_path,
        counts,
    )
    return counts


class PatchCheckerApp(ttk.Frame):
    def __init__(self, master: tk.Tk) -> None:
        super().__init__(master, padding=14)
        self.master = master
        self.old_var = tk.StringVar()
        self.current_var = tk.StringVar()
        self.output_var = tk.StringVar()
        self.status_var = tk.StringVar(value="Select both reports, then click Check patches.")
        self._build()

    def _build(self) -> None:
        self.grid(sticky="nsew")
        self.master.columnconfigure(0, weight=1)
        self.master.rowconfigure(0, weight=1)
        self.columnconfigure(1, weight=1)

        ttk.Label(self, text="Older missing-patches report:").grid(row=0, column=0, sticky="w", pady=5)
        ttk.Entry(self, textvariable=self.old_var).grid(row=0, column=1, sticky="ew", padx=8)
        ttk.Button(self, text="Browse…", command=self.browse_old).grid(row=0, column=2)

        ttk.Label(self, text="Current patch report:").grid(row=1, column=0, sticky="w", pady=5)
        ttk.Entry(self, textvariable=self.current_var).grid(row=1, column=1, sticky="ew", padx=8)
        ttk.Button(self, text="Browse…", command=self.browse_current).grid(row=1, column=2)

        ttk.Label(self, text="Output Excel file:").grid(row=2, column=0, sticky="w", pady=5)
        ttk.Entry(self, textvariable=self.output_var).grid(row=2, column=1, sticky="ew", padx=8)
        ttk.Button(self, text="Save as…", command=self.browse_output).grid(row=2, column=2)

        ttk.Separator(self).grid(row=3, column=0, columnspan=3, sticky="ew", pady=12)
        ttk.Button(self, text="Check patches", command=self.run_check).grid(row=4, column=0, columnspan=3, pady=4)
        ttk.Label(self, textvariable=self.status_var, wraplength=680).grid(row=5, column=0, columnspan=3, sticky="w", pady=(10, 0))

    @staticmethod
    def report_filetypes() -> list[tuple[str, str]]:
        return [("Patch reports", "*.xlsx *.xlsm *.csv"), ("Excel files", "*.xlsx *.xlsm"), ("CSV files", "*.csv"), ("All files", "*.*")]

    def browse_old(self) -> None:
        path = filedialog.askopenfilename(title="Select older missing-patches report", filetypes=self.report_filetypes())
        if path:
            self.old_var.set(path)
            if not self.output_var.get():
                source = Path(path)
                self.output_var.set(str(source.with_name(source.stem + "_checked.xlsx")))

    def browse_current(self) -> None:
        path = filedialog.askopenfilename(title="Select current patch report", filetypes=self.report_filetypes())
        if path:
            self.current_var.set(path)

    def browse_output(self) -> None:
        path = filedialog.asksaveasfilename(
            title="Save checked report",
            defaultextension=".xlsx",
            filetypes=[("Excel workbook", "*.xlsx")],
        )
        if path:
            self.output_var.set(path)

    def run_check(self) -> None:
        old_path = self.old_var.get().strip()
        current_path = self.current_var.get().strip()
        output_path = self.output_var.get().strip()

        if not old_path or not current_path or not output_path:
            messagebox.showwarning("Files required", "Select the older report, current report, and output file.")
            return

        self.status_var.set("Checking reports…")
        self.update_idletasks()

        try:
            counts = process_reports(old_path, current_path, output_path)
        except Exception as exc:
            self.status_var.set("Check failed.")
            details = "".join(traceback.format_exception_only(type(exc), exc)).strip()
            messagebox.showerror("Patch check failed", details)
            return

        installed = counts.get("NOW INSTALLED", 0)
        reboot = counts.get("INSTALLED - REBOOT REQUIRED", 0)
        outstanding = counts.get("STILL OUTSTANDING", 0)
        not_found = counts.get("NOT FOUND IN CURRENT REPORT", 0)
        message = (
            f"Finished. Installed: {installed}; reboot required: {reboot}; "
            f"outstanding: {outstanding}; not found: {not_found}.\n\nSaved to:\n{output_path}"
        )
        self.status_var.set(message.replace("\n\n", " "))
        messagebox.showinfo("Patch check complete", message)


def main() -> None:
    root = tk.Tk()
    root.title("Missing Patch Status Checker")
    root.geometry("760x250")
    root.minsize(650, 230)
    PatchCheckerApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
