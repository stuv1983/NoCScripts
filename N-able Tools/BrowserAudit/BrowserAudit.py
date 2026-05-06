#!/usr/bin/env python3
"""
Browser CVE Audit Report Generator v6
====================================
1. Click "Select Task Report CSV(s)" and pick your CSV file(s)
2. Click "Select Device Inventory" and pick the XLSX
3. Click "Generate Report"

Install dependencies once:
    pip install pandas openpyxl xlrd
"""

import re
import threading
import tkinter as tk
from datetime import datetime
from pathlib import Path
from tkinter import filedialog, messagebox, ttk

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ── Excel helpers ─────────────────────────────────────────────────────────────
COL = {
    "dark": "2C3E50", "navy": "1A5276", "red": "C0392B",
    "amber": "E67E22", "green": "1E8449", "gold": "7D6608",
    "row_orange": "FDEBD0", "row_blue": "EBF5FB",
    "row_gold": "FEF9E7", "row_gold2": "F9F3D2", "white": "FDFEFE",
}

def _fill(h):
    return PatternFill("solid", start_color=h)

HDR_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Arial", size=9)
WRAP      = Alignment(wrap_text=True, vertical="top")
CENTER    = Alignment(horizontal="center", vertical="center")
_thin     = Side(style="thin", color="BDBDBD")
BORDER    = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

def _hdr(ws, row, ncols, fill):
    for c in range(1, ncols + 1):
        cell = ws.cell(row, c)
        cell.fill = fill; cell.font = HDR_FONT; cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 22

def _widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Browser parsing ───────────────────────────────────────────────────────────
# N-able output can wrap/truncate the InstallScope column, for example:
#   P..., Per, Per-User, User, S..., Sys, System
# It can also contain paths with spaces such as "Program Files (x86)".
# So parse from Browser -> path ending in .exe -> version -> architecture -> optional scope.
_BPAT = re.compile(
    r"(Google Chrome|Microsoft Edge|Mozilla Firefox|Brave|Opera)"
    r"\s+(C:\\.*?\.exe)"
    r"\s+([\d.]+)"
    r"\s+(32-bit|64-bit)"
    r"(?:\s+([A-Za-z.\-]+))?",
    re.I,
)

def _normalise_scope(scope, path):
    s = (scope or "").strip().lower()
    p = (path or "").lower()

    # AppData path is the strongest signal for per-user install, even when
    # N-able truncates InstallScope to P... or omits the full value.
    if "\\appdata\\" in p or "appdata" in p:
        return "Per-User"
    if s.startswith(("p", "per", "user")):
        return "Per-User"
    if s.startswith(("s", "sys")):
        return "System"
    return "Unknown"

def parse_browsers(output):
    results = []
    text = str(output).replace("\r", "\n").replace(";", "\n")
    for m in _BPAT.finditer(text):
        path = m.group(2).strip()
        scope = _normalise_scope(m.group(5), path)
        results.append({
            "browser": m.group(1).strip(),
            "path": path,
            "version": m.group(3).strip(),
            "arch": m.group(4).strip(),
            "scope": scope,
        })
    return results

def detect_issues(browsers):
    by_name = {}
    for b in browsers:
        by_name.setdefault(b["browser"], []).append(b)
    dual = {n: e for n, e in by_name.items()
            if {"32-bit", "64-bit"}.issubset({x["arch"] for x in e})}
    per_user = [b for b in browsers if "AppData" in b["path"] or b["scope"] == "Per-User"]
    return dual, per_user


# ── Core logic ────────────────────────────────────────────────────────────────
def load_inventory(path, stale_days):
    inv = pd.read_excel(path)
    inv.columns = inv.columns.str.strip()
    inv["Device name"] = inv["Device name"].astype(str).str.strip()
    inv["last_response_dt"] = pd.to_datetime(
        inv["Last response (Local time)"], format="%m/%d/%y %I:%M:%S %p", errors="coerce")
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    inv["days_since"] = (today - inv["last_response_dt"]).dt.days
    stale = inv["days_since"] > stale_days
    return inv[~stale].copy(), inv[stale].copy()

def load_audit(paths):
    combined = pd.concat([pd.read_csv(p) for p in paths], ignore_index=True)
    combined.columns = combined.columns.str.strip()

    required = {"Task", "Status", "Device", "Date", "Output", "Client", "Site"}
    missing = required - set(combined.columns)
    if missing:
        raise ValueError(f"Task report CSV is missing required column(s): {', '.join(sorted(missing))}")

    combined["Device"] = combined["Device"].astype(str).str.strip()
    combined["Task"] = combined["Task"].astype(str).str.strip()
    combined["Status"] = combined["Status"].astype(str).str.strip()

    audit = combined[
        combined["Task"].str.contains("BrowserAudit|SimpleBrowserAudit", case=False, na=False) &
        combined["Status"].str.casefold().eq("stopped")
    ].copy()

    if audit.empty:
        raise ValueError("No stopped BrowserAudit/SimpleBrowserAudit task rows were found in the selected CSV.")

    # N-able exports dates as text such as '6 May 2026 14:51'. Parse where possible so the latest
    # result per device is selected correctly instead of relying on string sorting.
    audit["_date_dt"] = pd.to_datetime(audit["Date"], errors="coerce", dayfirst=True)
    audit = audit.sort_values(["_date_dt", "Date"], na_position="first").drop_duplicates(subset="Device", keep="last")
    audit = audit.drop(columns=["_date_dt"])
    return audit

def build_flagged(audit, inv_active):
    flagged = []
    idx = inv_active.set_index("Device name")
    for _, row in audit.iterrows():
        browsers = parse_browsers(str(row["Output"]))
        dual, per_user = detect_issues(browsers)
        issues, detail = [], []
        if dual:
            issues.append("Dual Install (32/64-bit)")
            for name, ents in dual.items():
                detail.append(name + ": " + " | ".join(f"{e['arch']} v{e['version']}" for e in ents))
        if per_user:
            issues.append("Per-User (AppData) Install")
            for b in per_user:
                detail.append(f"{b['browser']} v{b['version']} @ {b['path']}")
        if not issues:
            continue
        ir = idx.loc[row["Device"]] if row["Device"] in idx.index else None
        flagged.append({
            "Device": row["Device"],
            "Client": ir["Customer name"] if ir is not None and "Customer name" in ir.index else row.get("Client", ""),
            "Site": ir["Site name"] if ir is not None and "Site name" in ir.index else row.get("Site", ""),
            "Issue Type": " + ".join(issues), "Issue Detail": " | ".join(detail),
            "OS":            ir["OS version"]                    if ir is not None else "N/A",
            "Model":         ir["Model"]                         if ir is not None else "N/A",
            "Username":      ir["Username"]                      if ir is not None else "N/A",
            "Last Response": str(ir["Last response (Local time)"]) if ir is not None else "N/A",
        })
    return flagged


def _browser_summary(installs, browser):
    rows = [b for b in installs if b["browser"] == browser]
    if not rows:
        return ""
    return " | ".join(
        f"{b['version']} {b['arch']} {b['scope']} @ {b['path']}"
        for b in rows
    )

def build_browser_matrix(audit_active, inv_active):
    """One row per active scanned device showing Chrome/Firefox/Edge presence and risk flags."""
    idx = inv_active.set_index("Device name")
    rows = []
    browsers_to_track = ["Google Chrome", "Mozilla Firefox", "Microsoft Edge"]

    for _, row in audit_active.iterrows():
        device = row["Device"]
        inv = idx.loc[device] if device in idx.index else None
        all_installs = parse_browsers(str(row["Output"]))
        installs = [b for b in all_installs if b["browser"] in browsers_to_track]

        rec = {
            "Device": device,
            "Client": inv["Customer name"] if inv is not None and "Customer name" in inv.index else row.get("Client", ""),
            "Site": inv["Site name"] if inv is not None and "Site name" in inv.index else row.get("Site", ""),
            "OS": inv["OS version"] if inv is not None else "N/A",
            "Username": inv["Username"] if inv is not None else "N/A",
            "Last Response": str(inv["Last response (Local time)"]) if inv is not None else "N/A",
        }
        for browser in browsers_to_track:
            browser_rows = [b for b in installs if b["browser"] == browser]
            rec[f"Has {browser}"] = "Yes" if browser_rows else "No"
            rec[f"{browser} Install Count"] = len(browser_rows)
            rec[f"{browser} Details"] = _browser_summary(installs, browser)
            rec[f"{browser} 32-bit"] = "Yes" if any(b["arch"] == "32-bit" for b in browser_rows) else "No"
            rec[f"{browser} Per-User"] = "Yes" if any(b["scope"] == "Per-User" or "appdata" in b["path"].lower() for b in browser_rows) else "No"

        rec["Any 32-bit Browser"] = "Yes" if any(b["arch"] == "32-bit" for b in all_installs) else "No"
        rec["Any Per-User Browser"] = "Yes" if any(b["scope"] == "Per-User" or "appdata" in b["path"].lower() for b in all_installs) else "No"
        rec["Total Browser Installs"] = len(all_installs)
        rows.append(rec)

    return rows

def build_browser_installs(audit_active, inv_active):
    """One row per detected Chrome/Firefox/Edge install for filtering down to 32-bit/per-user installs."""
    idx = inv_active.set_index("Device name")
    rows = []
    browsers_to_track = {"Google Chrome", "Mozilla Firefox", "Microsoft Edge"}

    for _, row in audit_active.iterrows():
        device = row["Device"]
        inv = idx.loc[device] if device in idx.index else None
        for b in parse_browsers(str(row["Output"])):
            if b["browser"] not in browsers_to_track:
                continue
            is_per_user = b["scope"] == "Per-User" or "appdata" in b["path"].lower()
            is_32bit = b["arch"] == "32-bit"
            rows.append({
                "Device": device,
                "Client": inv["Customer name"] if inv is not None and "Customer name" in inv.index else row.get("Client", ""),
                "Site": inv["Site name"] if inv is not None and "Site name" in inv.index else row.get("Site", ""),
                "Browser": b["browser"],
                "Version": b["version"],
                "Architecture": b["arch"],
                "Install Scope": b["scope"],
                "Install Path": b["path"],
                "Is 32-bit": "Yes" if is_32bit else "No",
                "Is Per-User/AppData": "Yes" if is_per_user else "No",
                "OS": inv["OS version"] if inv is not None else "N/A",
                "Username": inv["Username"] if inv is not None else "N/A",
                "Last Response": str(inv["Last response (Local time)"]) if inv is not None else "N/A",
            })
    return rows

def browser_overview_counts(browser_matrix, browser_installs):
    browsers = ["Google Chrome", "Mozilla Firefox", "Microsoft Edge"]
    counts = []
    for browser in browsers:
        devices = {r["Device"] for r in browser_matrix if r[f"Has {browser}"] == "Yes"}
        devices_32 = {r["Device"] for r in browser_installs if r["Browser"] == browser and r["Is 32-bit"] == "Yes"}
        devices_per = {r["Device"] for r in browser_installs if r["Browser"] == browser and r["Is Per-User/AppData"] == "Yes"}
        installs = [r for r in browser_installs if r["Browser"] == browser]
        counts.append({
            "Browser": browser,
            "Devices with Browser": len(devices),
            "Total Installs Detected": len(installs),
            "Devices with 32-bit Install": len(devices_32),
            "Devices with Per-User/AppData Install": len(devices_per),
        })
    return counts

def build_report(flagged, not_scanned, stale_inv, inv_active, audit_active, browser_matrix, browser_installs, output_path, stale_days):
    wb = Workbook()

    ws = wb.active; ws.title = "Summary"; ws.sheet_view.showGridLines = False
    ws["A1"] = "Browser CVE Audit Report"
    ws["A1"].font = Font(name="Arial", bold=True, size=16, color=COL["dark"])
    ws["A2"] = f"Generated {datetime.now().strftime('%d %b %Y')}  |  Devices inactive >{stale_days} days excluded"
    ws["A2"].font = Font(name="Arial", size=10, color="7F8C8D")
    for i, (label, val, color) in enumerate([
        ("Total Fleet (Inventory)",              len(inv_active) + len(stale_inv), COL["dark"]),
        (f"Stale / Offline (>{stale_days} days)", len(stale_inv),                  COL["gold"]),
        ("Active Devices (Inventory)",           len(inv_active),                  COL["navy"]),
        ("Active Devices Scanned",               len(audit_active),                COL["green"]),
        ("Devices with Issues",                  len(flagged),                     COL["red"]),
        ("Active Devices NOT Scanned",           len(not_scanned),                 COL["amber"]),
    ], 4):
        f = _fill(color)
        ws.cell(i, 1, label).font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        ws.cell(i, 1).fill = f; ws.cell(i, 1).border = BORDER
        ws.cell(i, 1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.cell(i, 2, val).font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
        ws.cell(i, 2).fill = f; ws.cell(i, 2).border = BORDER; ws.cell(i, 2).alignment = CENTER
        ws.row_dimensions[i].height = 22
    # Browser overview counts
    browser_counts = browser_overview_counts(browser_matrix, browser_installs)
    start_row = 12
    headers = ["Browser", "Devices with Browser", "Total Installs Detected", "Devices with 32-bit Install", "Devices with Per-User/AppData Install"]
    for c, h in enumerate(headers, 1):
        ws.cell(start_row, c, h)
    _hdr(ws, start_row, len(headers), _fill(COL["navy"]))
    for r, rec in enumerate(browser_counts, start_row + 1):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(r, c, rec[h]); cell.font = BODY_FONT
            cell.border = BORDER; cell.alignment = CENTER if c > 1 else WRAP
        ws.row_dimensions[r].height = 22

    # Tracking totals for quick triage
    tracking_row = start_row + len(browser_counts) + 3
    any_32 = len({r["Device"] for r in browser_installs if r["Is 32-bit"] == "Yes"})
    any_per = len({r["Device"] for r in browser_installs if r["Is Per-User/AppData"] == "Yes"})
    total_installs = len(browser_installs)
    for i, (label, val, color) in enumerate([
        ("Devices with ANY 32-bit browser install", any_32, COL["amber"]),
        ("Devices with ANY per-user/AppData install", any_per, COL["red"]),
        ("Total Chrome/Firefox/Edge installs detected", total_installs, COL["green"]),
    ], tracking_row):
        f = _fill(color)
        ws.cell(i, 1, label).font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        ws.cell(i, 1).fill = f; ws.cell(i, 1).border = BORDER
        ws.cell(i, 1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.cell(i, 2, val).font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
        ws.cell(i, 2).fill = f; ws.cell(i, 2).border = BORDER; ws.cell(i, 2).alignment = CENTER
        ws.row_dimensions[i].height = 22

    _widths(ws, [42, 18, 22, 24, 32])

    ws2 = wb.create_sheet("⚠ Devices with Issues"); ws2.sheet_view.showGridLines = False
    hdrs = ["Device", "Client", "Site", "Issue Type", "Issue Detail", "OS", "Model", "Username", "Last Response"]
    for c, h in enumerate(hdrs, 1): ws2.cell(1, c, h)
    _hdr(ws2, 1, len(hdrs), _fill(COL["red"]))
    for r, rec in enumerate(flagged, 2):
        alt = _fill(COL["row_orange"]) if r % 2 == 0 else _fill(COL["white"])
        for c, key in enumerate(hdrs, 1):
            cell = ws2.cell(r, c, rec.get(key, "")); cell.font = BODY_FONT
            cell.border = BORDER; cell.fill = alt; cell.alignment = WRAP
        ws2.row_dimensions[r].height = 45
    _widths(ws2, [20, 14, 24, 26, 58, 36, 30, 26, 20]); ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = ws2.dimensions

    ws_browser = wb.create_sheet("🌐 Browser Devices"); ws_browser.sheet_view.showGridLines = False
    br_h = [
        "Device", "Client", "Site", "OS", "Username", "Last Response",
        "Has Google Chrome", "Google Chrome Install Count", "Google Chrome Details", "Google Chrome 32-bit", "Google Chrome Per-User",
        "Has Mozilla Firefox", "Mozilla Firefox Install Count", "Mozilla Firefox Details", "Mozilla Firefox 32-bit", "Mozilla Firefox Per-User",
        "Has Microsoft Edge", "Microsoft Edge Install Count", "Microsoft Edge Details", "Microsoft Edge 32-bit", "Microsoft Edge Per-User",
        "Any 32-bit Browser", "Any Per-User Browser", "Total Browser Installs",
    ]
    for c, h in enumerate(br_h, 1): ws_browser.cell(1, c, h)
    _hdr(ws_browser, 1, len(br_h), _fill(COL["navy"]))
    for r, rec in enumerate(browser_matrix, 2):
        is_risk = rec["Any 32-bit Browser"] == "Yes" or rec["Any Per-User Browser"] == "Yes"
        alt = _fill(COL["row_orange"]) if is_risk else (_fill(COL["row_blue"]) if r % 2 == 0 else _fill(COL["white"]))
        for c, key in enumerate(br_h, 1):
            cell = ws_browser.cell(r, c, rec.get(key, "")); cell.font = BODY_FONT
            cell.border = BORDER; cell.fill = alt; cell.alignment = WRAP
        ws_browser.row_dimensions[r].height = 22
    _widths(ws_browser, [18, 14, 22, 32, 24, 20, 14, 16, 60, 14, 16, 14, 16, 60, 14, 16, 14, 16, 60, 14, 16, 16, 18, 14])
    ws_browser.freeze_panes = "A2"
    ws_browser.auto_filter.ref = ws_browser.dimensions

    ws_inst = wb.create_sheet("📦 Browser Installs"); ws_inst.sheet_view.showGridLines = False
    inst_h = ["Device", "Client", "Site", "Browser", "Version", "Architecture", "Install Scope", "Install Path", "Is 32-bit", "Is Per-User/AppData", "OS", "Username", "Last Response"]
    for c, h in enumerate(inst_h, 1): ws_inst.cell(1, c, h)
    _hdr(ws_inst, 1, len(inst_h), _fill(COL["green"]))
    for r, rec in enumerate(browser_installs, 2):
        is_risk = rec["Is 32-bit"] == "Yes" or rec["Is Per-User/AppData"] == "Yes"
        alt = _fill(COL["row_orange"]) if is_risk else (_fill(COL["white"]) if r % 2 else _fill(COL["row_blue"]))
        for c, key in enumerate(inst_h, 1):
            cell = ws_inst.cell(r, c, rec.get(key, "")); cell.font = BODY_FONT
            cell.border = BORDER; cell.fill = alt; cell.alignment = WRAP
        ws_inst.row_dimensions[r].height = 24
    _widths(ws_inst, [18, 14, 22, 20, 18, 14, 16, 68, 12, 20, 32, 24, 20])
    ws_inst.freeze_panes = "A2"
    ws_inst.auto_filter.ref = ws_inst.dimensions

    ws3 = wb.create_sheet("🔍 Not Scanned (Active)"); ws3.sheet_view.showGridLines = False
    ns_h = ["Device Name", "Customer", "Site", "Device Type", "OS Version", "Username", "Last Response", "Manufacturer", "Model"]
    for c, h in enumerate(ns_h, 1): ws3.cell(1, c, h)
    _hdr(ws3, 1, len(ns_h), _fill(COL["amber"]))
    for r, (_, rec) in enumerate(not_scanned.iterrows(), 2):
        alt = _fill(COL["row_blue"]) if r % 2 == 0 else _fill(COL["white"])
        vals = [rec["Device name"], rec["Customer name"], rec["Site name"], rec["Device type"],
                rec["OS version"], rec["Username"], str(rec["Last response (Local time)"]), rec["Manufacturer"], rec["Model"]]
        for c, v in enumerate(vals, 1):
            cell = ws3.cell(r, c, v); cell.font = BODY_FONT
            cell.border = BORDER; cell.fill = alt; cell.alignment = WRAP
        ws3.row_dimensions[r].height = 18
    _widths(ws3, [22, 14, 24, 14, 38, 26, 20, 16, 34]); ws3.freeze_panes = "A2"

    ws4 = wb.create_sheet(f"🕐 Stale >{stale_days}d (Excluded)"); ws4.sheet_view.showGridLines = False
    st_h = ["Device Name", "Customer", "Site", "Device Type", "OS Version", "Username", "Last Response", "Days Offline", "Manufacturer", "Model"]
    for c, h in enumerate(st_h, 1): ws4.cell(1, c, h)
    _hdr(ws4, 1, len(st_h), _fill(COL["gold"]))
    for r, (_, rec) in enumerate(stale_inv.sort_values("days_since", ascending=False).iterrows(), 2):
        alt = _fill(COL["row_gold2"]) if r % 2 == 0 else _fill(COL["row_gold"])
        days = int(rec["days_since"]) if pd.notna(rec["days_since"]) else "?"
        vals = [rec["Device name"], rec["Customer name"], rec["Site name"], rec["Device type"],
                rec["OS version"], rec["Username"], str(rec["Last response (Local time)"]), days, rec["Manufacturer"], rec["Model"]]
        for c, v in enumerate(vals, 1):
            cell = ws4.cell(r, c, v); cell.font = BODY_FONT
            cell.border = BORDER; cell.fill = alt; cell.alignment = WRAP
        ws4.row_dimensions[r].height = 18
    _widths(ws4, [22, 14, 24, 14, 38, 26, 20, 18, 16, 34]); ws4.freeze_panes = "A2"

    wb.save(output_path)
    return len(flagged), len(not_scanned), len(stale_inv)



def get_output_path():
    """Return a writable report path. Handles redirected/missing Desktop folders."""
    filename = f"Browser_CVE_Report_{datetime.now().strftime('%Y-%m-%d_%H%M%S')}.xlsx"
    candidates = [
        Path.home() / "Desktop",
        Path.home() / "OneDrive" / "Desktop",
        Path.home() / "Downloads",
        Path.cwd(),
    ]
    for folder in candidates:
        try:
            folder.mkdir(parents=True, exist_ok=True)
            test_file = folder / ".browser_cve_write_test"
            test_file.write_text("test", encoding="utf-8")
            test_file.unlink(missing_ok=True)
            return folder / filename
        except Exception:
            continue
    return Path.cwd() / filename


# ═══════════════════════════════════════════════════════════════════════════════
#  GUI — three big buttons, a log, and a run button
# ═══════════════════════════════════════════════════════════════════════════════
class App(tk.Tk):
    BG      = "#1E2530"
    CARD    = "#252D3A"
    ACCENT  = "#3B82F6"
    ACCENTH = "#2563EB"
    TEXT    = "#E2E8F0"
    SUBTEXT = "#94A3B8"
    SUCCESS = "#22C55E"
    ERROR   = "#EF4444"
    WARN    = "#F59E0B"
    BORDERC = "#374151"

    def __init__(self):
        super().__init__()
        self.title("Browser CVE Audit v6")
        self.configure(bg=self.BG)
        self.resizable(False, False)

        self.task_files = []
        self.inv_path   = ""
        self.out_path   = ""
        self.stale_days = tk.IntVar(value=30)

        self._build()
        w, h = 540, 550
        self.update_idletasks()
        x = (self.winfo_screenwidth()  - w) // 2
        y = (self.winfo_screenheight() - h) // 2
        self.geometry(f"{w}x{h}+{x}+{y}")

    # ── helpers ───────────────────────────────────────────────────────────────
    def _lbl(self, p, text, size=10, bold=False, color=None):
        return tk.Label(p, text=text,
                        font=("Segoe UI", size, "bold" if bold else "normal"),
                        fg=color or self.TEXT, bg=p.cget("bg"))

    def _big_btn(self, parent, text, sub, cmd, done_var):
        """Card-style button with a status sub-label."""
        card = tk.Frame(parent, bg=self.CARD, cursor="hand2",
                        highlightthickness=1, highlightbackground=self.BORDERC)
        card.pack(fill="x", padx=20, pady=6)
        inner = tk.Frame(card, bg=self.CARD)
        inner.pack(fill="x", padx=14, pady=10)
        tk.Label(inner, text=text, font=("Segoe UI", 10, "bold"),
                 fg=self.TEXT, bg=self.CARD, anchor="w").pack(fill="x")
        status = tk.Label(inner, textvariable=done_var, font=("Segoe UI", 8),
                          fg=self.SUBTEXT, bg=self.CARD, anchor="w")
        status.pack(fill="x")
        for widget in (card, inner, status):
            widget.bind("<Button-1>", lambda e: cmd())
            widget.bind("<Enter>",    lambda e: card.config(highlightbackground=self.ACCENT))
            widget.bind("<Leave>",    lambda e: card.config(highlightbackground=self.BORDERC))
        return card

    # ── layout ────────────────────────────────────────────────────────────────
    def _build(self):
        # Header
        hdr = tk.Frame(self, bg="#141920", pady=16)
        hdr.pack(fill="x")
        self._lbl(hdr, "🔍  Browser CVE Audit", size=15, bold=True).pack()
        self._lbl(hdr, "Select input files, choose where to save, then generate",
                  size=9, color=self.SUBTEXT).pack(pady=(2, 0))

        # Step labels
        steps = tk.Frame(self, bg=self.BG)
        steps.pack(fill="x", padx=20, pady=(14, 2))
        self._lbl(steps, "STEP 1 & 2 — Select input files", size=8,
                  color=self.SUBTEXT).pack(anchor="w")

        # CSV button
        self.csv_label = tk.StringVar(value="No files selected")
        self._big_btn(self, "📄  Task Report CSV(s)", "",
                      self._pick_csvs, self.csv_label)

        # Inventory button
        self.inv_label = tk.StringVar(value="No file selected")
        self._big_btn(self, "📋  Device Inventory XLSX", "",
                      self._pick_inv, self.inv_label)

        # Output file button
        self.out_label = tk.StringVar(value="No save location selected")
        self._big_btn(self, "💾  Save Report As", "",
                      self._pick_output, self.out_label)

        # Stale days row
        stale_row = tk.Frame(self, bg=self.BG)
        stale_row.pack(fill="x", padx=20, pady=(4, 0))
        self._lbl(stale_row, "Exclude devices offline for more than", size=9).pack(side="left")
        tk.Spinbox(stale_row, from_=1, to=365, textvariable=self.stale_days,
                   width=4, font=("Segoe UI", 9), bg="#2D3748", fg=self.TEXT,
                   buttonbackground="#374151", relief="flat",
                   highlightthickness=1, highlightbackground=self.BORDERC).pack(
            side="left", padx=6)
        self._lbl(stale_row, "days", size=9).pack(side="left")

        # Log
        log_frame = tk.Frame(self, bg=self.BG)
        log_frame.pack(fill="x", padx=20, pady=(12, 0))
        self._lbl(log_frame, "STEP 3 — Run", size=8, color=self.SUBTEXT).pack(anchor="w")
        self.log = tk.Text(log_frame, height=5, font=("Consolas", 9),
                           bg="#0F1419", fg="#A8B2C1", relief="flat",
                           state="disabled", wrap="word",
                           highlightthickness=1, highlightbackground=self.BORDERC)
        self.log.pack(fill="x", pady=(4, 0))
        self.log.tag_config("ok",   foreground=self.SUCCESS)
        self.log.tag_config("err",  foreground=self.ERROR)
        self.log.tag_config("warn", foreground=self.WARN)
        self.log.tag_config("dim",  foreground=self.SUBTEXT)

        # Generate button
        foot = tk.Frame(self, bg=self.BG, pady=14)
        foot.pack(fill="x")
        self.run_btn = tk.Button(
            foot, text="▶  Generate Report", command=self._run,
            font=("Segoe UI", 11, "bold"), bg=self.ACCENT, fg=self.TEXT,
            activebackground=self.ACCENTH, activeforeground=self.TEXT,
            relief="flat", cursor="hand2", padx=20, pady=8)
        self.run_btn.pack()
        self.run_btn.bind("<Enter>", lambda e: self.run_btn.config(bg=self.ACCENTH))
        self.run_btn.bind("<Leave>", lambda e: self.run_btn.config(bg=self.ACCENT))
        self.progress = ttk.Progressbar(foot, mode="indeterminate", length=260)
        self.progress.pack(pady=(8, 0))

    # ── file pickers ──────────────────────────────────────────────────────────
    def _pick_csvs(self):
        files = filedialog.askopenfilenames(
            title="Select Task Report CSV(s)",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")])
        if files:
            self.task_files = list(files)
            names = ", ".join(Path(f).name for f in files)
            self.csv_label.set(f"✔  {len(files)} file(s): {names}")

    def _pick_inv(self):
        f = filedialog.askopenfilename(
            title="Select Device Inventory XLSX",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All", "*.*")])
        if f:
            self.inv_path = f
            self.inv_label.set(f"✔  {Path(f).name}")

    def _pick_output(self):
        default_name = f"Browser_CVE_Report_{datetime.now().strftime('%Y-%m-%d_%H%M%S')}.xlsx"
        f = filedialog.asksaveasfilename(
            title="Choose where to save the Browser CVE report",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel workbook", "*.xlsx"), ("All files", "*.*")],
        )
        if f:
            if not f.lower().endswith(".xlsx"):
                f += ".xlsx"
            self.out_path = f
            self.out_label.set(f"✔  {f}")

    # ── logging ───────────────────────────────────────────────────────────────
    def _log(self, msg, tag=""):
        self.log.config(state="normal")
        self.log.insert("end", msg + "\n", tag)
        self.log.see("end")
        self.log.config(state="disabled")

    # ── run ───────────────────────────────────────────────────────────────────
    def _run(self):
        if not self.task_files:
            messagebox.showerror("Missing", "Please select at least one Task Report CSV.")
            return
        if not self.inv_path:
            messagebox.showerror("Missing", "Please select the Device Inventory XLSX.")
            return
        if not self.out_path:
            self._pick_output()
            if not self.out_path:
                messagebox.showerror("Missing", "Please choose where to save the report.")
                return

        self.log.config(state="normal"); self.log.delete("1.0", "end"); self.log.config(state="disabled")
        self.run_btn.config(state="disabled")
        self.progress.start(12)
        threading.Thread(target=self._worker, daemon=True).start()

    def _worker(self):
        try:
            stale_days = self.stale_days.get()
            out_path = self.out_path

            self._log(f"Task CSV(s): {len(self.task_files)} selected", "dim")
            for file in self.task_files:
                self._log(f"  CSV: {file}", "dim")
            self._log(f"Inventory: {self.inv_path}", "dim")
            self._log(f"Save as: {out_path}", "dim")

            self._log("Loading inventory…", "dim")
            inv_active, stale_inv = load_inventory(self.inv_path, stale_days)
            self._log(f"  Active: {len(inv_active)}  |  Stale: {len(stale_inv)}", "dim")

            self._log(f"Loading {len(self.task_files)} CSV(s)…", "dim")
            audit = load_audit(self.task_files)
            self._log(f"  Audit records: {len(audit)}", "dim")

            # Inventory is the source of truth for the selected client/scope.
            # Only devices present in the active inventory are included in scanned/issue/browser counts.
            active_names = set(inv_active["Device name"])
            audit_device_names = set(audit["Device"])
            audit_active = audit[audit["Device"].isin(active_names)].copy()
            scanned_names = set(audit_active["Device"])
            not_scanned = inv_active[
                inv_active["Device name"].isin(active_names - scanned_names)
            ].sort_values(["Site name", "Device name"])
            ignored_not_inventory = sorted(audit_device_names - active_names)
            if ignored_not_inventory:
                self._log(f"  Ignored scanned CSV devices not in inventory: {len(ignored_not_inventory)}", "warn")
            self._log(f"  In-scope scanned devices: {len(audit_active)}", "dim")

            self._log("Checking browser installs…", "dim")
            flagged = build_flagged(audit_active, inv_active)
            browser_matrix = build_browser_matrix(audit_active, inv_active)
            browser_installs = build_browser_installs(audit_active, inv_active)

            self._log("Writing report…", "dim")
            Path(out_path).parent.mkdir(parents=True, exist_ok=True)
            n_f, n_ns, n_st = build_report(
                flagged, not_scanned, stale_inv, inv_active, audit_active,
                browser_matrix, browser_installs,
                out_path, stale_days)

            self._log(f"\n✔  Saved to: {out_path}", "ok")
            self._log(f"   Issues found    : {n_f}",  "warn" if n_f  else "ok")
            self._log(f"   Browser installs: {len(browser_installs)}", "dim")
            self._log(f"   32-bit devices  : {len({r['Device'] for r in browser_installs if r['Is 32-bit'] == 'Yes'})}", "warn")
            self._log(f"   Per-user devices: {len({r['Device'] for r in browser_installs if r['Is Per-User/AppData'] == 'Yes'})}", "warn")
            self._log(f"   Not scanned     : {n_ns}", "warn" if n_ns else "ok")
            self._log(f"   Stale excluded  : {n_st}", "dim")

            self.after(0, lambda: self._offer_open(out_path))

        except Exception as exc:
            import traceback
            self._log(f"\n✘  {exc}", "err")
            self._log(traceback.format_exc(), "err")
        finally:
            self.after(0, self._done)

    def _done(self):
        self.progress.stop()
        self.run_btn.config(state="normal")

    def _offer_open(self, path):
        if messagebox.askyesno("Done!", f"Report saved to:\n{path}\n\nOpen it now?"):
            import os, subprocess, sys
            if os.name == "nt":
                os.startfile(path)
            elif sys.platform == "darwin":
                subprocess.Popen(["open", path])
            else:
                subprocess.Popen(["xdg-open", path])


if __name__ == "__main__":
    App().mainloop()
