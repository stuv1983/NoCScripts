"""Patch Report Comparison Tool.

Compares two N-able patch overview exports and writes a highlighted copy of the
second (current) report so you can see at a glance what is Installed, Installing,
Missing, Pending or Reboot Required - and which rows are new since the previous
report.

The main focus is follow-up: every patch that was Failed or Missing in report 1
is tracked through to its status in report 2 and listed on its own sheet, so you
can see what was fixed, what is still outstanding and what dropped off the report.

Rows are matched between reports on Device + Patch name.

Requires: openpyxl (pip install openpyxl)
"""

from __future__ import annotations

import collections
import datetime as dt
import html
import os
import queue
import re
import subprocess
import sys
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover - surfaced to the user in the GUI
    openpyxl = None


APP_TITLE = "Patch Report Comparison"

# ---------------------------------------------------------------------------
# Highlight styles.  Colours mirror Excel's built-in cell styles but are written
# as explicit fills so they render the same regardless of the workbook theme.
# A fill of None means font colour only (Excel's "Warning Text" behaves this way).
# ---------------------------------------------------------------------------


class Hl:
    """A highlight: display label plus fill and font colours."""

    def __init__(self, label, fill, font, bold=False):
        self.label = label
        self.fill_hex = fill
        self.font_hex = font
        self.bold = bold
        self._fill = None
        self._font = None

    @property
    def fill(self):
        if self._fill is None and self.fill_hex:
            self._fill = PatternFill("solid", start_color=self.fill_hex,
                                     end_color=self.fill_hex)
        return self._fill

    @property
    def font(self):
        if self._font is None:
            self._font = Font(color=self.font_hex, bold=self.bold)
        return self._font

    def apply(self, cell):
        if self.fill is not None:
            cell.fill = self.fill
        cell.font = self.font

    def swatch_colours(self):
        """(background, foreground) for the tkinter preview swatch."""
        return "#" + (self.fill_hex or "FFFFFF"), "#" + self.font_hex


def build_styles():
    return {
        "Installed": Hl("Good - green", "C6EFCE", "006100"),
        "Installing": Hl("Note - yellow", "FFEB9C", "9C6500"),
        "Missing": Hl("Bad - red", "FFC7CE", "9C0006"),
        "Reboot Required": Hl("Accent1 - light blue", "BDD7EE", "1F4E79"),
        "Failed": Hl("Accent2 - orange", "F8CBAD", "833C0C"),
        "Pending": Hl("Warning Text - red text", None, "C00000"),
        "Ignored": Hl("Accent5 - teal", "4BACC6", "FFFFFF"),
        "__new__": Hl("Accent3 - green", "A9D08E", "375623"),
    }


# Statuses the user can choose to highlight, in display order.
HIGHLIGHT_STATUSES = ["Installed", "Installing", "Missing", "Reboot Required",
                      "Failed", "Pending", "Ignored"]

# Statuses in report 1 that are tracked through to report 2.
FOCUS_STATUSES = ("Failed", "Missing")

FOLLOWUP_SHEET = "Failed and Missing"

# Change classifications written to the "Change" column.
CHG_NEW_DEVICE = "New device"
CHG_NEW_PATCH = "New patch"
CHG_CHANGED = "Status changed"
CHG_UNCHANGED = "Unchanged"

# Follow-up outcomes for patches that were Failed or Missing in report 1.
OUT_RESOLVED = "Resolved"
OUT_IN_PROGRESS = "In progress"
OUT_OUTSTANDING = "Outstanding"
OUT_DROPPED = "No longer reported"
OUTCOMES = (OUT_OUTSTANDING, OUT_IN_PROGRESS, OUT_RESOLVED, OUT_DROPPED)

OUTCOME_STYLE = {
    OUT_RESOLVED: "Installed",       # green
    OUT_IN_PROGRESS: "Installing",   # yellow
    OUT_OUTSTANDING: "Missing",      # red
    OUT_DROPPED: "Reboot Required",  # light blue
}

# Report-2 statuses that count as still-broken / in-flight / fixed.
STILL_BROKEN = {"failed", "missing"}
IN_FLIGHT = {"pending", "installing", "reboot required"}

REQUIRED_COLUMNS = ["Client", "Site", "Device", "Status", "Patch"]
DATE_COLUMN = "Discovered / Install Date"

# Clients that are ticked for exclusion the first time the list is built.
DEFAULT_EXCLUDED_CLIENTS = ("Pop-Up Health", "SEN", "Energy Power Systems")

# Device inventory report (optional third input).
INV_DEVICE = "device name"
INV_CLIENT = "customer name"
INV_SITE = "site name"
INV_TYPE = "device type"
INV_OS = "os version"
INV_AGE = "device age"
INV_LAST_SEEN = ("last response (local time)", "last response (utc)", "last response")
INV_REQUIRED = (INV_DEVICE, INV_CLIENT)

# Output sheet holding one row per device.
DEVICES_SHEET = "Devices"

# Devices that need chasing: no RMM record at all, or an agent that has stopped
# checking in.  Written to their own sheet so they can be worked as a list.
MISSING_SHEET = "Missing from RMM"
WHY_NO_RMM = "Not in RMM inventory"
WHY_NEVER = "Never responded"
WHY_STALE = "No recent response"
WHY_ORDER = (WHY_NO_RMM, WHY_NEVER, WHY_STALE)
WHY_STYLE = {
    WHY_NO_RMM: "Missing",     # red - patched by something we cannot see
    WHY_NEVER: "Missing",      # red - in the RMM but has never called home
    WHY_STALE: "Installing",   # yellow - was fine, has gone quiet
}

# Where a device turned up.
SEEN_BOTH = "In both reports"
SEEN_ONLY_1 = "Report 1 only"
SEEN_ONLY_2 = "Report 2 only"
SEEN_ONLY_INV = "No patch data"
SEEN_ORDER = (SEEN_ONLY_1, SEEN_ONLY_INV, SEEN_ONLY_2, SEEN_BOTH)
SEEN_STYLE = {
    SEEN_ONLY_1: "Missing",       # red - stopped appearing in the patch report
    SEEN_ONLY_INV: "Installing",  # yellow - known device, no patch rows at all
    SEEN_ONLY_2: "__new__",       # green - new since report 1
    SEEN_BOTH: None,
}

# Days since last check-in before a device is flagged.  The upper figure is the
# default for the "not seen for" box in the GUI and can be changed per run.
STALE_WARN_DAYS = 7
STALE_BAD_DAYS = 30

FILTER_ALL = "all"
FILTER_CHANGED = "changed"
FILTER_FOCUS = "focus"


# ---------------------------------------------------------------------------
# Data loading / comparison
# ---------------------------------------------------------------------------


def clean(value):
    """Normalise a cell value for comparison (HTML entities, case, whitespace)."""
    if value is None:
        return ""
    text = html.unescape(str(value))
    return re.sub(r"\s+", " ", text).strip().lower()


def display(value):
    """Tidy a cell value for output (report 1 exports are HTML-escaped)."""
    if isinstance(value, str):
        return html.unescape(value).strip()
    return value


def find_header_row(ws, max_scan=10):
    """Return (row_index, {lowercase_name: column_index}) for the header row."""
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        if row is None:
            continue
        lookup = {}
        for col_idx, cell in enumerate(row):
            if cell is None:
                continue
            lookup[re.sub(r"\s+", " ", str(cell)).strip().lower()] = col_idx
        if all(name.lower() in lookup for name in REQUIRED_COLUMNS):
            return row_idx, lookup
    return None, None


def load_report(path, log):
    """Read a patch report into (rows, column_lookup)."""
    log(f"Reading {os.path.basename(path)} ...")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        header_row, lookup = find_header_row(ws)
        if lookup is None:
            raise ValueError(
                f"{os.path.basename(path)}: could not find a header row containing "
                + ", ".join(REQUIRED_COLUMNS)
            )
        rows = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if row is None or all(cell is None for cell in row):
                continue
            rows.append(row)
        title = ws.title
    finally:
        wb.close()
    log(f"  {len(rows):,} data rows, sheet '{title}'")
    return rows, lookup


def parse_when(value):
    """Best-effort parse of an inventory timestamp into a datetime."""
    if isinstance(value, dt.datetime):
        return value
    if isinstance(value, dt.date):
        return dt.datetime.combine(value, dt.time())
    if not value:
        return None
    text = str(value).replace(" ", " ").replace("\xa0", " ").strip()
    if not text or text.lower() in ("never", "n/a", "-"):
        return None
    for fmt in ("%m/%d/%y %I:%M:%S %p", "%m/%d/%Y %I:%M:%S %p",
                "%d/%m/%Y %I:%M:%S %p", "%Y-%m-%d %H:%M:%S",
                "%m/%d/%y %H:%M:%S", "%m/%d/%y", "%m/%d/%Y", "%Y-%m-%d"):
        try:
            return dt.datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def load_inventory(path, log):
    """Read a device inventory export into {clean device name: {field: value}}."""
    log(f"Reading {os.path.basename(path)} ...")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        header_row = None
        lookup = {}
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
            if row is None:
                continue
            candidate = {}
            for col_idx, cell in enumerate(row):
                if cell is None:
                    continue
                candidate[re.sub(r"\s+", " ", str(cell)).strip().lower()] = col_idx
            if all(name in candidate for name in INV_REQUIRED):
                header_row, lookup = row_idx, candidate
                break
        if header_row is None:
            raise ValueError(
                f"{os.path.basename(path)}: could not find a header row containing "
                "'Device name' and 'Customer name'"
            )

        last_seen_col = next((lookup[name] for name in INV_LAST_SEEN if name in lookup), None)

        def get(row, name):
            idx = lookup.get(name)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        devices = {}
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if row is None or all(cell is None for cell in row):
                continue
            key = clean(get(row, INV_DEVICE))
            if not key:
                continue
            last_seen = row[last_seen_col] if (last_seen_col is not None
                                               and last_seen_col < len(row)) else None
            devices[key] = {
                "client": display(get(row, INV_CLIENT)),
                "site": display(get(row, INV_SITE)),
                "device": display(get(row, INV_DEVICE)),
                "type": display(get(row, INV_TYPE)),
                "os": display(get(row, INV_OS)),
                "age": display(get(row, INV_AGE)),
                "last_seen": parse_when(last_seen),
                # Kept so an unparseable value ("Never", an odd date format) is
                # shown as-is rather than as an empty cell.
                "last_seen_text": display(last_seen),
            }
    finally:
        wb.close()
    log(f"  {len(devices):,} devices in inventory")
    return devices


def cell_at(row, lookup, name):
    idx = lookup.get(name.lower())
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def row_key(row, lookup):
    return (clean(cell_at(row, lookup, "Device")), clean(cell_at(row, lookup, "Patch")))


def index_previous(rows, lookup):
    """Build the lookups needed to compare report 2 against report 1.

    Returns (statuses_by_key, devices, focus_by_key) where focus_by_key holds the
    report-1 rows whose status was Failed or Missing.
    """
    statuses = {}
    devices = set()
    focus = {}
    for row in rows:
        key = row_key(row, lookup)
        status = str(cell_at(row, lookup, "Status") or "").strip()
        devices.add(key[0])
        statuses.setdefault(key, set()).add(status)
        if status in FOCUS_STATUSES:
            focus.setdefault(key, {"row": row, "statuses": set()})["statuses"].add(status)
    return statuses, devices, focus


def classify(status, key, prev_statuses, prev_devices):
    """Return (change, previous_status_text) for a current-report row."""
    previous = prev_statuses.get(key)
    if previous is None:
        if key[0] not in prev_devices:
            return CHG_NEW_DEVICE, ""
        return CHG_NEW_PATCH, ""
    prev_text = " / ".join(sorted(s for s in previous if s))
    if status in previous:
        return CHG_UNCHANGED, prev_text
    return CHG_CHANGED, prev_text


def scan_clients(paths, inventory_path=None, log=None):
    """Distinct client names across the given inputs, for the exclusion list.

    The inventory is included so a client that only has devices there - no patch
    rows at all - can still be ticked out of the device sheet.
    """
    noop = log or (lambda _message: None)
    names = {}

    def add(value):
        if value:
            names.setdefault(clean(value), value)

    for path in paths:
        if not path or not os.path.isfile(path):
            continue
        rows, lookup = load_report(path, noop)
        for row in rows:
            add(display(cell_at(row, lookup, "Client")))
    if inventory_path and os.path.isfile(inventory_path):
        for info in load_inventory(inventory_path, noop).values():
            add(info["client"])
    return [names[key] for key in sorted(names)]


def drop_excluded(rows, lookup, excluded):
    """Remove rows belonging to an excluded client."""
    if not excluded:
        return rows, 0
    kept = [row for row in rows if clean(cell_at(row, lookup, "Client")) not in excluded]
    return kept, len(rows) - len(kept)


def outcome_for(status):
    """Follow-up outcome for a patch that was Failed or Missing in report 1."""
    normalised = status.strip().lower()
    if normalised in STILL_BROKEN:
        return OUT_OUTSTANDING
    if normalised in IN_FLIGHT:
        return OUT_IN_PROGRESS
    return OUT_RESOLVED


# ---------------------------------------------------------------------------
# Output workbook
# ---------------------------------------------------------------------------

def _style_header(ws, columns, widths):
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="44546A", end_color="44546A")
    for col_idx, name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"


def _finish_sheet(ws, columns, last_row):
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{max(last_row, 1)}"


def collect_devices(prev_rows, prev_lookup, curr_rows, curr_lookup, prev_focus, inventory):
    """One record per device across both reports and the inventory."""
    devices = {}

    def record(key, client=None, site=None):
        entry = devices.get(key)
        if entry is None:
            entry = devices[key] = {
                "client": "", "site": "", "device": "",
                "counts": {}, "in1": False, "in2": False,
                "was_bad": 0, OUT_OUTSTANDING: 0, OUT_IN_PROGRESS: 0,
                OUT_RESOLVED: 0, OUT_DROPPED: 0,
            }
        if client and not entry["client"]:
            entry["client"] = client
        if site and not entry["site"]:
            entry["site"] = site
        return entry

    for row in prev_rows:
        key = clean(cell_at(row, prev_lookup, "Device"))
        entry = record(key, display(cell_at(row, prev_lookup, "Client")),
                       display(cell_at(row, prev_lookup, "Site")))
        entry["in1"] = True
        entry["device"] = entry["device"] or display(cell_at(row, prev_lookup, "Device"))

    # Report 2 wins for the client / site / device casing shown in the output.
    current_status = {}
    for row in curr_rows:
        key = clean(cell_at(row, curr_lookup, "Device"))
        entry = record(key)
        entry["in2"] = True
        entry["client"] = display(cell_at(row, curr_lookup, "Client")) or entry["client"]
        entry["site"] = display(cell_at(row, curr_lookup, "Site")) or entry["site"]
        entry["device"] = display(cell_at(row, curr_lookup, "Device")) or entry["device"]
        status = str(cell_at(row, curr_lookup, "Status") or "").strip()
        entry["counts"][status] = entry["counts"].get(status, 0) + 1
        current_status[row_key(row, curr_lookup)] = status

    for key in prev_focus:
        entry = record(key[0])
        entry["was_bad"] += 1
        status = current_status.get(key)
        outcome = OUT_DROPPED if status is None else outcome_for(status)
        entry[outcome] += 1

    for key, info in inventory.items():
        entry = record(key, info["client"], info["site"])
        entry["device"] = entry["device"] or info["device"]
        entry["inventory"] = info

    now = dt.datetime.now()
    for key, entry in devices.items():
        info = entry.setdefault("inventory", None) or {}
        entry["device"] = entry["device"] or key
        entry["last_seen"] = info.get("last_seen")
        entry["last_seen_text"] = info.get("last_seen_text")
        entry["days"] = (now - entry["last_seen"]).days if entry["last_seen"] else None
        if entry["in1"] and entry["in2"]:
            entry["seen"] = SEEN_BOTH
        elif entry["in1"]:
            entry["seen"] = SEEN_ONLY_1
        elif entry["in2"]:
            entry["seen"] = SEEN_ONLY_2
        else:
            entry["seen"] = SEEN_ONLY_INV
    return devices


def _write_devices(wb, devices, styles, stale_days):
    ws = wb.create_sheet(DEVICES_SHEET)
    counted = ["Installed", "Pending", "Installing", "Reboot Required", "Failed",
               "Missing", "Ignored"]
    # The four outcome columns break "Was Failed/Missing" down in full, so the
    # row adds up: was_bad == outstanding + in progress + resolved + dropped.
    columns = (["Client", "Site", "Device", "Seen in", "Type", "OS", "Last Response",
                "Days Since", "Device Age"] + counted +
               ["Patches", "Was Failed/Missing"] + list(OUTCOMES))
    widths = ([26, 20, 24, 16, 14, 44, 20, 12, 12] + [11] * len(counted) +
              [10, 18, 13, 12, 11, 19])
    _style_header(ws, columns, widths)
    col = {name: i for i, name in enumerate(columns, start=1)}

    order = {name: i for i, name in enumerate(SEEN_ORDER)}
    rows = sorted(devices.values(),
                  key=lambda e: (order.get(e["seen"], 99), str(e["client"]).lower(),
                                 str(e["device"]).lower()))

    for out_row, entry in enumerate(rows, start=2):
        info = entry["inventory"] or {}
        days = entry["days"]
        total_patches = sum(entry["counts"].values())
        values = ([entry["client"], entry["site"], entry["device"], entry["seen"],
                   info.get("type"), info.get("os"),
                   entry["last_seen"] or entry["last_seen_text"], days, info.get("age")] +
                  [entry["counts"].get(name, 0) for name in counted] +
                  [total_patches, entry["was_bad"]] +
                  [entry[outcome] for outcome in OUTCOMES])
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=out_row, column=col_idx, value=value)
            if isinstance(value, (dt.datetime, dt.date)):
                cell.number_format = "dd-mmm-yyyy hh:mm"

        style_name = SEEN_STYLE[entry["seen"]]
        if style_name:
            styles[style_name].apply(ws.cell(row=out_row, column=col["Seen in"]))

        if days is None:
            # The inventory gave a last response we could not turn into a date -
            # "Never", or a format we do not know.  Either way it needs a flag.
            # A device with nothing in the field at all is left alone, so an
            # export without that column does not come out solid red.
            if entry["last_seen_text"]:
                styles["Missing"].apply(ws.cell(row=out_row, column=col["Last Response"]))
        elif days > STALE_WARN_DAYS:
            style = styles["Missing"] if days > stale_days else styles["Installing"]
            style.apply(ws.cell(row=out_row, column=col["Days Since"]))

        for name in ("Failed", "Missing"):
            if entry["counts"].get(name):
                styles[name].apply(ws.cell(row=out_row, column=col[name]))
        for outcome in OUTCOMES:
            if entry[outcome]:
                styles[OUTCOME_STYLE[outcome]].apply(
                    ws.cell(row=out_row, column=col[outcome]))

    _finish_sheet(ws, columns, len(rows) + 1)
    return collections.Counter(entry["seen"] for entry in devices.values())


def why_missing(entry, stale_days):
    """Why a device needs chasing, or None if it is checking in normally.

    Only devices carrying patch rows can be "not in the RMM inventory" - a device
    that exists solely in the inventory is by definition in the RMM.
    """
    if entry["inventory"] is None:
        return WHY_NO_RMM
    if entry["days"] is None:
        # "Never", or a date we could not read.  A blank field tells us nothing,
        # so it is not treated as a fault.
        return WHY_NEVER if entry["last_seen_text"] else None
    return WHY_STALE if entry["days"] > stale_days else None


def _write_missing(wb, devices, styles, stale_days):
    """Sheet of devices with no RMM record or a silent agent."""
    ws = wb.create_sheet(MISSING_SHEET)
    columns = ["Client", "Site", "Device", "Why", "Last Response", "Days Since",
               "Seen in", "Type", "OS", "Device Age", "Patches",
               "Was Failed/Missing", OUT_OUTSTANDING]
    _style_header(ws, columns, [26, 20, 24, 20, 20, 12, 16, 14, 40, 12, 10, 18, 13])
    col = {name: i for i, name in enumerate(columns, start=1)}

    order = {name: i for i, name in enumerate(WHY_ORDER)}
    flagged = []
    for entry in devices.values():
        why = why_missing(entry, stale_days)
        if why:
            flagged.append((why, entry))
    # Worst first: no RMM record, then never seen, then longest silent.
    flagged.sort(key=lambda pair: (order.get(pair[0], 99), -(pair[1]["days"] or 0),
                                   str(pair[1]["client"]).lower(),
                                   str(pair[1]["device"]).lower()))

    for out_row, (why, entry) in enumerate(flagged, start=2):
        info = entry["inventory"] or {}
        values = [entry["client"], entry["site"], entry["device"], why,
                  entry["last_seen"] or entry["last_seen_text"], entry["days"],
                  entry["seen"], info.get("type"), info.get("os"), info.get("age"),
                  sum(entry["counts"].values()), entry["was_bad"],
                  entry[OUT_OUTSTANDING]]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=out_row, column=col_idx, value=value)
            if isinstance(value, (dt.datetime, dt.date)):
                cell.number_format = "dd-mmm-yyyy hh:mm"
        styles[WHY_STYLE[why]].apply(ws.cell(row=out_row, column=col["Why"]))
        if entry["days"] is not None:
            styles[WHY_STYLE[why]].apply(ws.cell(row=out_row, column=col["Days Since"]))
        if entry[OUT_OUTSTANDING]:
            styles["Missing"].apply(ws.cell(row=out_row, column=col[OUT_OUTSTANDING]))

    _finish_sheet(ws, columns, len(flagged) + 1)
    return collections.Counter(why for why, _entry in flagged)


def write_output(path, rows, lookup, prev_statuses, prev_devices, prev_focus,
                 prev_lookup, prev_rows, inventory, options, log, progress):
    styles = build_styles()
    new_style = styles["__new__"]
    selected = {s for s in HIGHLIGHT_STATUSES if options["statuses"].get(s)}

    columns = ["Client", "Site", "Device", "Status", "Patch", DATE_COLUMN,
               "Previous Status", "Change"]
    widths = [26, 20, 24, 16, 62, 20, 18, 16]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Patch Report"
    _style_header(ws, columns, widths)

    status_counts = {}
    change_counts = {}
    transitions = {}
    outcome_counts = {}
    followup_rows = []   # rows for the follow-up sheet
    seen_focus = set()
    written = 0
    total = len(rows)
    out_row = 2

    for i, row in enumerate(rows):
        if i % 500 == 0:
            progress(i, total)

        key = row_key(row, lookup)
        status = str(cell_at(row, lookup, "Status") or "").strip()
        change, prev_text = classify(status, key, prev_statuses, prev_devices)
        is_new = change in (CHG_NEW_DEVICE, CHG_NEW_PATCH)

        status_counts[status or "(blank)"] = status_counts.get(status or "(blank)", 0) + 1
        change_counts[change] = change_counts.get(change, 0) + 1
        if change == CHG_CHANGED:
            label = f"{prev_text} -> {status}"
            transitions[label] = transitions.get(label, 0) + 1

        focus = prev_focus.get(key)
        outcome = ""
        if focus is not None:
            seen_focus.add(key)
            outcome = outcome_for(status)
            outcome_counts[outcome] = outcome_counts.get(outcome, 0) + 1
            was = " / ".join(sorted(focus["statuses"]))
            followup_rows.append((outcome, [
                display(cell_at(row, lookup, "Client")),
                display(cell_at(row, lookup, "Site")),
                display(cell_at(row, lookup, "Device")),
                display(cell_at(row, lookup, "Patch")),
                was,
                status,
                cell_at(row, lookup, DATE_COLUMN),
            ]))

        if options["filter_mode"] == FILTER_CHANGED and change == CHG_UNCHANGED:
            continue
        if options["filter_mode"] == FILTER_FOCUS and focus is None:
            continue

        values = [
            display(cell_at(row, lookup, "Client")),
            display(cell_at(row, lookup, "Site")),
            display(cell_at(row, lookup, "Device")),
            status,
            display(cell_at(row, lookup, "Patch")),
            cell_at(row, lookup, DATE_COLUMN),
            prev_text,
            change,
        ]

        status_hl = styles.get(status) if status in selected else None
        row_hl = None
        if is_new and options["highlight_new"] and options["new_fills_row"]:
            row_hl = new_style
        elif status_hl is not None and options["whole_row"]:
            row_hl = status_hl

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=out_row, column=col_idx, value=value)
            if isinstance(value, (dt.datetime, dt.date)):
                cell.number_format = "dd-mmm-yyyy"
            if row_hl is not None:
                row_hl.apply(cell)

        # When the row is not filled wholesale, the Status cell still carries
        # its own colour.
        if status_hl is not None and row_hl is None:
            status_hl.apply(ws.cell(row=out_row, column=4))

        # New rows are flagged in the Change column even when the row is filled
        # by status, so nothing new can hide behind a status colour.
        if is_new and options["highlight_new"] and not options["new_fills_row"]:
            new_style.apply(ws.cell(row=out_row, column=8))

        out_row += 1
        written += 1

    progress(total, total)
    _finish_sheet(ws, columns, out_row - 1)

    # Anything that was Failed or Missing in report 1 and is absent from report 2.
    # These are almost always superseded versions - the old Chrome / Defender
    # definition disappears once a newer one is offered - so they are counted in
    # the run log but kept off the follow-up sheet unless explicitly asked for.
    for key, focus in prev_focus.items():
        if key in seen_focus:
            continue
        row = focus["row"]
        outcome_counts[OUT_DROPPED] = outcome_counts.get(OUT_DROPPED, 0) + 1
        if not options["include_dropped"]:
            continue
        followup_rows.append((OUT_DROPPED, [
            display(cell_at(row, prev_lookup, "Client")),
            display(cell_at(row, prev_lookup, "Site")),
            display(cell_at(row, prev_lookup, "Device")),
            display(cell_at(row, prev_lookup, "Patch")),
            " / ".join(sorted(focus["statuses"])),
            "",
            None,
        ]))

    _write_followup(wb, followup_rows, styles)

    log("Building the device report ...")
    stale_days = options["stale_days"]
    devices = collect_devices(prev_rows, prev_lookup, rows, lookup, prev_focus, inventory)
    seen_counts = _write_devices(wb, devices, styles, stale_days)
    log(f"  {len(devices):,} devices across both reports and the inventory")

    # Without an inventory there is nothing to be missing from, so the sheet is
    # skipped rather than listing every device as absent from the RMM.
    if inventory:
        why_counts = _write_missing(wb, devices, styles, stale_days)
        log(f"  {sum(why_counts.values()):,} devices missing from the RMM or not "
            f"responding within {stale_days} days")
    else:
        why_counts = collections.Counter()
        log(f"  no device inventory supplied - skipping the '{MISSING_SHEET}' sheet")

    # Summary sheet turned off - the same figures are reported in the run log.
    # _write_summary(wb, status_counts, change_counts, transitions, outcome_counts,
    #                selected, options, styles, total, written, len(prev_focus),
    #                seen_counts, why_counts, bool(inventory), devices)

    wb.save(path)
    log(f"Saved {written:,} rows to {path}")
    return {
        "status_counts": status_counts,
        "change_counts": change_counts,
        "transitions": transitions,
        "outcome_counts": outcome_counts,
        "focus_total": len(prev_focus),
        "written": written,
        "seen_counts": seen_counts,
        "device_total": len(devices),
        "why_counts": why_counts,
        "had_inventory": bool(inventory),
    }


def _write_followup(wb, followup_rows, styles):
    ws = wb.create_sheet(FOLLOWUP_SHEET)
    columns = ["Client", "Site", "Device", "Patch", "Report 1 Status",
               "Report 2 Status", "Discovered / Install Date"]
    _style_header(ws, columns, [26, 20, 24, 62, 16, 18, 22])

    order = {name: i for i, name in enumerate(OUTCOMES)}
    followup_rows.sort(
        key=lambda r: (order.get(r[0], 99), str(r[1][0]), str(r[1][2]), str(r[1][3])))

    for out_row, (outcome, values) in enumerate(followup_rows, start=2):
        # The whole row is coloured by its report 2 status.  Rows that dropped
        # off the report have no status, so they take the outcome's colour.
        style = styles.get(values[5]) or styles[OUTCOME_STYLE[outcome]]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=out_row, column=col_idx, value=value)
            if isinstance(value, (dt.datetime, dt.date)):
                cell.number_format = "dd-mmm-yyyy"
            style.apply(cell)

    _finish_sheet(ws, columns, len(followup_rows) + 1)


def _write_summary(wb, status_counts, change_counts, transitions, outcome_counts,
                   selected, options, styles, total, written, focus_total,
                   seen_counts, why_counts, had_inventory, devices):
    ws = wb.create_sheet("Summary")
    bold = Font(bold=True)
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 12

    def heading(row, text):
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = Font(bold=True, size=12)
        return row + 1

    r = 1
    r = heading(r, "Failed / Missing in report 1 - follow-up")
    ws.cell(row=r, column=1, value="Tracked patches").font = bold
    ws.cell(row=r, column=2, value=focus_total).font = bold
    r += 1
    for outcome in OUTCOMES:
        cell = ws.cell(row=r, column=1, value=outcome)
        if outcome != OUT_DROPPED or options["include_dropped"]:
            styles[OUTCOME_STYLE[outcome]].apply(cell)
        ws.cell(row=r, column=2, value=outcome_counts.get(outcome, 0))
        r += 1
    if not options["include_dropped"]:
        ws.cell(row=r, column=1, value=(
            "'No longer reported' is counted here only - the patch is absent from "
            "report 2, nearly always because a newer version superseded it."))
        r += 1
    ws.cell(row=r, column=1, value=f"See the '{FOLLOWUP_SHEET}' sheet for the detail.")
    r += 2

    r = heading(r, "Devices")
    ws.cell(row=r, column=1, value="Devices seen").font = bold
    ws.cell(row=r, column=2, value=sum(seen_counts.values())).font = bold
    r += 1
    for seen in SEEN_ORDER:
        cell = ws.cell(row=r, column=1, value=seen)
        style_name = SEEN_STYLE[seen]
        if style_name:
            styles[style_name].apply(cell)
        ws.cell(row=r, column=2, value=seen_counts.get(seen, 0))
        r += 1
    ws.cell(row=r, column=1, value=f"See the '{DEVICES_SHEET}' sheet for the detail.")
    r += 2

    if had_inventory:
        stale_days = options["stale_days"]
        responding = sum(1 for e in devices.values()
                         if why_missing(e, stale_days) is None)
        r = heading(r, "Device check-in (last response)")
        ws.cell(row=r, column=1, value=f"Responding within {stale_days} days").font = bold
        ws.cell(row=r, column=2, value=responding).font = bold
        r += 1
        for why in WHY_ORDER:
            label = (f"{why} ({stale_days}+ days)" if why == WHY_STALE else why)
            styles[WHY_STYLE[why]].apply(ws.cell(row=r, column=1, value=label))
            ws.cell(row=r, column=2, value=why_counts.get(why, 0))
            r += 1
        ws.cell(row=r, column=1, value=f"See the '{MISSING_SHEET}' sheet for the detail.")
        r += 2

    r = heading(r, "Legend")
    for status in HIGHLIGHT_STATUSES:
        style = styles[status]
        style.apply(ws.cell(row=r, column=1, value=status))
        note = style.label if status in selected else f"{style.label} - not highlighted"
        ws.cell(row=r, column=2, value=note)
        r += 1
    new_style = styles["__new__"]
    new_style.apply(ws.cell(row=r, column=1, value="New since previous report"))
    ws.cell(row=r, column=2,
            value=new_style.label if options["highlight_new"] else "Accent3 - not highlighted")
    r += 2

    r = heading(r, "Current status totals")
    ws.cell(row=r, column=1, value="Status").font = bold
    ws.cell(row=r, column=2, value="Rows").font = bold
    r += 1
    for status, count in sorted(status_counts.items(), key=lambda kv: -kv[1]):
        ws.cell(row=r, column=1, value=status)
        ws.cell(row=r, column=2, value=count)
        r += 1
    r += 1

    r = heading(r, "Change against previous report")
    for change in (CHG_UNCHANGED, CHG_CHANGED, CHG_NEW_PATCH, CHG_NEW_DEVICE):
        ws.cell(row=r, column=1, value=change)
        ws.cell(row=r, column=2, value=change_counts.get(change, 0))
        r += 1
    r += 1

    if transitions:
        r = heading(r, "Status transitions")
        ws.cell(row=r, column=1, value="From -> To").font = bold
        ws.cell(row=r, column=2, value="Rows").font = bold
        r += 1
        for label, count in sorted(transitions.items(), key=lambda kv: -kv[1]):
            ws.cell(row=r, column=1, value=label)
            ws.cell(row=r, column=2, value=count)
            r += 1
        r += 1

    filter_labels = {
        FILTER_ALL: "All rows",
        FILTER_CHANGED: "New or changed only",
        FILTER_FOCUS: "Failed / Missing in report 1 only",
    }
    r = heading(r, "Run details")
    for label, value in [
        ("Rows in current report", total),
        ("Rows written", written),
        ("Rows filtered", filter_labels[options["filter_mode"]]),
        ("Flag devices not seen for", f"{options['stale_days']} days"),
        ("Highlight scope", "Entire row" if options["whole_row"] else "Status cell only"),
        ("New rows", "Fill entire row" if options["new_fills_row"] else "Mark Change column"),
        ("Generated", dt.datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]:
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=value)
        r += 1


def run_comparison(previous_path, current_path, inventory_path, output_path, options,
                   log, progress):
    prev_rows, prev_lookup = load_report(previous_path, log)
    curr_rows, curr_lookup = load_report(current_path, log)
    inventory = load_inventory(inventory_path, log) if inventory_path else {}

    excluded = {clean(name) for name in options.get("excluded_clients", ())}
    if excluded:
        prev_rows, dropped1 = drop_excluded(prev_rows, prev_lookup, excluded)
        curr_rows, dropped2 = drop_excluded(curr_rows, curr_lookup, excluded)
        before = len(inventory)
        inventory = {key: info for key, info in inventory.items()
                     if clean(info["client"]) not in excluded}
        log(f"Excluding {len(excluded)} client(s): "
            f"{dropped1:,} + {dropped2:,} patch rows and "
            f"{before - len(inventory):,} inventory devices removed")

    log("Indexing previous report ...")
    prev_statuses, prev_devices, prev_focus = index_previous(prev_rows, prev_lookup)
    log(f"  {len(prev_statuses):,} device/patch pairs across {len(prev_devices):,} devices")
    log(f"  {len(prev_focus):,} patches were Failed or Missing")

    log("Comparing and writing output ...")
    return write_output(output_path, curr_rows, curr_lookup, prev_statuses, prev_devices,
                        prev_focus, prev_lookup, prev_rows, inventory, options,
                        log, progress)


# ---------------------------------------------------------------------------
# GUI
# ---------------------------------------------------------------------------


class PatchReportApp(ttk.Frame):
    def __init__(self, master):
        super().__init__(master, padding=12)
        self.master = master
        self.grid(row=0, column=0, sticky="nsew")
        master.columnconfigure(0, weight=1)
        master.rowconfigure(0, weight=1)
        self.columnconfigure(0, weight=1)

        self.previous_path = tk.StringVar()
        self.current_path = tk.StringVar()
        self.inventory_path = tk.StringVar()
        self.output_path = tk.StringVar()
        self.client_vars = {}
        self.client_note = tk.StringVar(value="Select both reports to list clients.")
        self.status_vars = {s: tk.BooleanVar(value=True) for s in HIGHLIGHT_STATUSES}
        self.highlight_new = tk.BooleanVar(value=True)
        self.new_fills_row = tk.BooleanVar(value=False)
        self.whole_row = tk.BooleanVar(value=True)
        self.include_dropped = tk.BooleanVar(value=False)
        self.stale_days = tk.StringVar(value=str(STALE_BAD_DAYS))
        self.filter_mode = tk.StringVar(value=FILTER_ALL)
        self.open_when_done = tk.BooleanVar(value=True)
        self.status_text = tk.StringVar(value="Select both reports to begin.")
        self.last_dir = os.path.expanduser("~")
        self.queue = queue.Queue()
        self.worker = None
        self.styles = build_styles() if openpyxl is not None else {}

        self._build_files()
        self._build_clients()
        self._build_options()
        self._build_log()
        self._build_actions()
        self.after(100, self._drain_queue)

    # -- layout ------------------------------------------------------------
    def _build_files(self):
        frame = ttk.LabelFrame(self, text="Reports", padding=10)
        frame.grid(row=0, column=0, sticky="ew")
        frame.columnconfigure(1, weight=1)

        rows = [
            ("Report 1 (previous):", self.previous_path, self._browse_previous),
            ("Report 2 (current):", self.current_path, self._browse_current),
            ("Device inventory (optional):", self.inventory_path, self._browse_inventory),
            ("Output file:", self.output_path, self._browse_output),
        ]
        for r, (label, var, command) in enumerate(rows):
            ttk.Label(frame, text=label).grid(row=r, column=0, sticky="w", pady=3)
            ttk.Entry(frame, textvariable=var).grid(row=r, column=1, sticky="ew", padx=6, pady=3)
            ttk.Button(frame, text="Browse...", command=command).grid(row=r, column=2, pady=3)

        ttk.Label(
            frame,
            text=("Report 2 is the one that gets highlighted. Rows match on Device + Patch. "
                  "Patches that were Failed or Missing in report 1 get their own sheet, and "
                  "the inventory adds last check-in and hardware detail to the device sheet."),
            foreground="#555555", wraplength=780, justify="left",
        ).grid(row=len(rows), column=0, columnspan=3, sticky="w", pady=(6, 0))

    def _build_clients(self):
        frame = ttk.LabelFrame(self, text="Exclude clients (ticked clients are left out)",
                               padding=10)
        frame.grid(row=1, column=0, sticky="ew", pady=(10, 0))
        frame.columnconfigure(0, weight=1)

        canvas = tk.Canvas(frame, height=104, highlightthickness=0,
                           background=self.winfo_toplevel().cget("background"))
        canvas.grid(row=0, column=0, sticky="ew")
        scroll = ttk.Scrollbar(frame, orient="vertical", command=canvas.yview)
        scroll.grid(row=0, column=1, sticky="ns")
        canvas.configure(yscrollcommand=scroll.set)

        self.client_frame = ttk.Frame(canvas)
        window = canvas.create_window((0, 0), window=self.client_frame, anchor="nw")
        self.client_frame.bind(
            "<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfigure(window, width=e.width))

        # Bound only while the pointer is over the list, so the wheel still
        # scrolls the log pane and the rest of the window normally.
        def wheel(event):
            canvas.yview_scroll(-1 * (event.delta // 120), "units")

        canvas.bind("<Enter>", lambda e: canvas.bind_all("<MouseWheel>", wheel))
        canvas.bind("<Leave>", lambda e: canvas.unbind_all("<MouseWheel>"))

        buttons = ttk.Frame(frame)
        buttons.grid(row=1, column=0, columnspan=2, sticky="w", pady=(6, 0))
        ttk.Button(buttons, text="Refresh list",
                   command=self._refresh_clients).grid(row=0, column=0)
        ttk.Button(buttons, text="Tick none",
                   command=lambda: self._set_all_clients(False)).grid(row=0, column=1, padx=6)
        ttk.Button(buttons, text="Tick all",
                   command=lambda: self._set_all_clients(True)).grid(row=0, column=2)
        ttk.Label(buttons, textvariable=self.client_note, foreground="#555555").grid(
            row=0, column=3, padx=(12, 0))

    def _set_all_clients(self, value):
        for var in self.client_vars.values():
            var.set(value)

    def _refresh_clients(self):
        previous, current = self.previous_path.get().strip(), self.current_path.get().strip()
        inventory = self.inventory_path.get().strip()
        if not (previous or current or inventory):
            return
        self.client_note.set("Reading client names ...")
        threading.Thread(target=self._scan_clients_worker,
                         args=(previous, current, inventory), daemon=True).start()

    def _scan_clients_worker(self, previous, current, inventory):
        try:
            names = scan_clients([previous, current], inventory)
        except Exception as exc:
            self.queue.put(("clients", ("error", str(exc))))
            return
        self.queue.put(("clients", ("ok", names)))

    def _populate_clients(self, names):
        for child in self.client_frame.winfo_children():
            child.destroy()
        previous = {clean(name): var.get() for name, var in self.client_vars.items()}
        defaults = {clean(name) for name in DEFAULT_EXCLUDED_CLIENTS}
        self.client_vars = {}
        for i, name in enumerate(names):
            key = clean(name)
            var = tk.BooleanVar(value=previous.get(key, key in defaults))
            self.client_vars[name] = var
            ttk.Checkbutton(self.client_frame, text=name, variable=var).grid(
                row=i // 3, column=i % 3, sticky="w", padx=(0, 18))
        ticked = sum(1 for var in self.client_vars.values() if var.get())
        self.client_note.set(f"{len(names)} clients, {ticked} excluded.")

    def _swatch(self, parent, style):
        bg, fg = style.swatch_colours()
        return tk.Label(parent, text=" Aa ", bg=bg, fg=fg, relief="solid",
                        borderwidth=1, font=("Segoe UI", 8))

    def _build_options(self):
        frame = ttk.LabelFrame(self, text="Highlighting", padding=10)
        frame.grid(row=2, column=0, sticky="ew", pady=(10, 0))
        frame.columnconfigure(0, weight=1)

        grid = ttk.Frame(frame)
        grid.grid(row=0, column=0, sticky="w")
        entries = [(s, self.styles[s].label if self.styles else "") for s in HIGHLIGHT_STATUSES]
        entries.append(("__new__", self.styles["__new__"].label if self.styles else ""))
        for i, (status, label) in enumerate(entries):
            r, c = i // 2, (i % 2) * 3
            text = "New since report 1" if status == "__new__" else status
            var = self.highlight_new if status == "__new__" else self.status_vars[status]
            if self.styles:
                self._swatch(grid, self.styles[status]).grid(row=r, column=c, padx=(0, 6), pady=2)
            ttk.Checkbutton(grid, text=text, variable=var).grid(
                row=r, column=c + 1, sticky="w")
            ttk.Label(grid, text=label, foreground="#777777").grid(
                row=r, column=c + 2, sticky="w", padx=(6, 28))

        ttk.Separator(frame, orient="horizontal").grid(row=1, column=0, sticky="ew", pady=8)

        opts = ttk.Frame(frame)
        opts.grid(row=2, column=0, sticky="w")
        ttk.Checkbutton(opts, text="Colour the entire row (otherwise the Status cell only)",
                        variable=self.whole_row).grid(row=0, column=0, sticky="w", pady=2)
        ttk.Checkbutton(opts, text="Fill new rows entirely with green (otherwise flag the Change column)",
                        variable=self.new_fills_row).grid(row=1, column=0, sticky="w", pady=2)
        ttk.Checkbutton(
            opts,
            text=("List patches that are no longer in report 2 on the '%s' sheet "
                  "(usually superseded versions)" % FOLLOWUP_SHEET),
            variable=self.include_dropped).grid(row=2, column=0, sticky="w", pady=2)
        ttk.Checkbutton(opts, text="Open the output file when finished",
                        variable=self.open_when_done).grid(row=3, column=0, sticky="w", pady=2)

        stale_frame = ttk.Frame(frame)
        stale_frame.grid(row=3, column=0, sticky="w", pady=(8, 0))
        ttk.Label(stale_frame, text="Treat a device as missing from the RMM after:").grid(
            row=0, column=0, sticky="w")
        ttk.Spinbox(stale_frame, from_=1, to=3650, width=6,
                    textvariable=self.stale_days).grid(row=0, column=1, padx=6)
        ttk.Label(stale_frame,
                  text=("days without a response. Those devices, plus any with no RMM "
                        "record at all, are listed on the '%s' sheet." % MISSING_SHEET),
                  foreground="#777777", wraplength=560, justify="left").grid(
            row=0, column=2, sticky="w")

        rows_frame = ttk.Frame(frame)
        rows_frame.grid(row=4, column=0, sticky="w", pady=(8, 0))
        ttk.Label(rows_frame, text="Rows to output:").grid(row=0, column=0, sticky="w", padx=(0, 10))
        for i, (mode, text) in enumerate([
            (FILTER_ALL, "All"),
            (FILTER_CHANGED, "New or changed only"),
            (FILTER_FOCUS, "Failed / Missing in report 1 only"),
        ]):
            ttk.Radiobutton(rows_frame, text=text, value=mode,
                            variable=self.filter_mode).grid(row=0, column=i + 1, sticky="w", padx=(0, 16))

    def _build_log(self):
        frame = ttk.LabelFrame(self, text="Progress", padding=10)
        frame.grid(row=3, column=0, sticky="nsew", pady=(10, 0))
        frame.columnconfigure(0, weight=1)
        frame.rowconfigure(1, weight=1)
        self.rowconfigure(3, weight=1)

        self.progress = ttk.Progressbar(frame, mode="determinate", maximum=100)
        self.progress.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 8))

        # Deliberately short: this is the only row that stretches, so the log
        # takes whatever height is left rather than forcing the window taller.
        self.log_widget = tk.Text(frame, height=8, wrap="none", state="disabled",
                                  font=("Consolas", 9))
        self.log_widget.grid(row=1, column=0, sticky="nsew")
        scroll = ttk.Scrollbar(frame, orient="vertical", command=self.log_widget.yview)
        scroll.grid(row=1, column=1, sticky="ns")
        self.log_widget.configure(yscrollcommand=scroll.set)

    def _build_actions(self):
        frame = ttk.Frame(self)
        frame.grid(row=4, column=0, sticky="ew", pady=(10, 0))
        frame.columnconfigure(0, weight=1)
        ttk.Label(frame, textvariable=self.status_text, foreground="#555555").grid(
            row=0, column=0, sticky="w")
        self.run_button = ttk.Button(frame, text="Compare && Highlight", command=self.start)
        self.run_button.grid(row=0, column=1, padx=(8, 0))
        ttk.Button(frame, text="Close", command=self.master.destroy).grid(row=0, column=2, padx=(8, 0))

    # -- file pickers ------------------------------------------------------
    def _ask_open(self, title):
        path = filedialog.askopenfilename(
            title=title, initialdir=self.last_dir,
            filetypes=[("Excel workbooks", "*.xlsx *.xlsm"), ("All files", "*.*")])
        if path:
            self.last_dir = os.path.dirname(path)
        return path

    def _browse_previous(self):
        path = self._ask_open("Select report 1 (previous)")
        if path:
            self.previous_path.set(path)
            self._suggest_output()
            self._refresh_clients()

    def _browse_current(self):
        path = self._ask_open("Select report 2 (current)")
        if path:
            self.current_path.set(path)
            self._suggest_output()
            self._refresh_clients()

    def _browse_inventory(self):
        path = self._ask_open("Select the device inventory report")
        if path:
            self.inventory_path.set(path)
            self._refresh_clients()

    def _browse_output(self):
        path = filedialog.asksaveasfilename(
            title="Save highlighted report as", initialdir=self.last_dir,
            defaultextension=".xlsx", filetypes=[("Excel workbook", "*.xlsx")])
        if path:
            self.output_path.set(path)

    def _suggest_output(self):
        current = self.current_path.get()
        if current and not self.output_path.get():
            base, _ = os.path.splitext(current)
            self.output_path.set(f"{base}_highlighted.xlsx")

    # -- worker plumbing ---------------------------------------------------
    def log(self, message):
        self.queue.put(("log", message))

    def _drain_queue(self):
        try:
            while True:
                kind, payload = self.queue.get_nowait()
                if kind == "log":
                    self.log_widget.configure(state="normal")
                    self.log_widget.insert("end", payload + "\n")
                    self.log_widget.see("end")
                    self.log_widget.configure(state="disabled")
                elif kind == "progress":
                    done, total = payload
                    self.progress["value"] = (done / total * 100) if total else 0
                elif kind == "status":
                    self.status_text.set(payload)
                elif kind == "clients":
                    outcome, value = payload
                    if outcome == "ok":
                        self._populate_clients(value)
                    else:
                        self.client_note.set(f"Could not read client names: {value}")
                elif kind == "done":
                    self._finish(payload)
        except queue.Empty:
            pass
        self.after(100, self._drain_queue)

    def start(self):
        if openpyxl is None:
            messagebox.showerror(APP_TITLE, "openpyxl is not installed.\n\nRun: pip install openpyxl")
            return
        previous, current, output = (self.previous_path.get().strip(),
                                     self.current_path.get().strip(),
                                     self.output_path.get().strip())
        inventory = self.inventory_path.get().strip()
        for label, path in (("Report 1", previous), ("Report 2", current)):
            if not path:
                messagebox.showwarning(APP_TITLE, f"Please select {label}.")
                return
            if not os.path.isfile(path):
                messagebox.showerror(APP_TITLE, f"{label} not found:\n{path}")
                return
        if inventory and not os.path.isfile(inventory):
            messagebox.showerror(APP_TITLE, f"Device inventory not found:\n{inventory}")
            return
        if not output:
            messagebox.showwarning(APP_TITLE, "Please choose an output file.")
            return
        if os.path.abspath(output) in (os.path.abspath(previous), os.path.abspath(current)):
            messagebox.showerror(APP_TITLE, "The output file must not overwrite a source report.")
            return
        try:
            stale_days = int(str(self.stale_days.get()).strip())
            if stale_days < 1:
                raise ValueError
        except ValueError:
            messagebox.showwarning(
                APP_TITLE, "'Treat a device as missing from the RMM after' needs a "
                           "whole number of days, 1 or more.")
            return

        options = {
            "statuses": {s: v.get() for s, v in self.status_vars.items()},
            "highlight_new": self.highlight_new.get(),
            "new_fills_row": self.new_fills_row.get(),
            "whole_row": self.whole_row.get(),
            "include_dropped": self.include_dropped.get(),
            "stale_days": stale_days,
            "filter_mode": self.filter_mode.get(),
            "excluded_clients": [name for name, var in self.client_vars.items() if var.get()],
        }

        self.run_button.configure(state="disabled")
        self.progress["value"] = 0
        self.log_widget.configure(state="normal")
        self.log_widget.delete("1.0", "end")
        self.log_widget.configure(state="disabled")
        self.queue.put(("status", "Working..."))

        self.worker = threading.Thread(
            target=self._work, args=(previous, current, inventory, output, options),
            daemon=True)
        self.worker.start()

    def _work(self, previous, current, inventory, output, options):
        try:
            result = run_comparison(
                previous, current, inventory, output, options, self.log,
                lambda done, total: self.queue.put(("progress", (done, total))))
            self.queue.put(("done", ("ok", output, result)))
        except PermissionError:
            self.queue.put(("done", ("error", output,
                                     "Could not write the output file - it may be open in Excel.")))
        except Exception as exc:  # surfaced in the GUI
            self.queue.put(("done", ("error", output, str(exc))))

    def _finish(self, payload):
        outcome, output, result = payload
        self.run_button.configure(state="normal")
        if outcome == "error":
            self.progress["value"] = 0
            self.status_text.set("Failed.")
            self.log(f"ERROR: {result}")
            messagebox.showerror(APP_TITLE, result)
            return

        self.progress["value"] = 100
        focus_total = result["focus_total"]
        outcome_counts = result["outcome_counts"]
        self.log("")
        self.log(f"Failed / Missing in report 1: {focus_total:,} patches")
        for name in OUTCOMES:
            note = ""
            if name == OUT_DROPPED and not self.include_dropped.get():
                note = "  (not listed - superseded or removed)"
            self.log(f"  {name:<20}{outcome_counts.get(name, 0):>8,}{note}")
        self.log("")
        self.log(f"Devices: {result['device_total']:,}")
        for name in SEEN_ORDER:
            self.log(f"  {name:<20}{result['seen_counts'].get(name, 0):>8,}")
        self.log("")
        if result["had_inventory"]:
            why_counts = result["why_counts"]
            self.log(f"Missing from the RMM: {sum(why_counts.values()):,} devices")
            for name in WHY_ORDER:
                self.log(f"  {name:<22}{why_counts.get(name, 0):>8,}")
        else:
            self.log("Missing from the RMM: no device inventory supplied.")
        self.log("")
        self.log("Current status totals:")
        for status, count in sorted(result["status_counts"].items(), key=lambda kv: -kv[1]):
            self.log(f"  {status:<20}{count:>8,}")
        self.log("")
        self.log("Against the previous report:")
        for change in (CHG_UNCHANGED, CHG_CHANGED, CHG_NEW_PATCH, CHG_NEW_DEVICE):
            self.log(f"  {change:<20}{result['change_counts'].get(change, 0):>8,}")
        transitions = result["transitions"]
        if transitions:
            self.log("")
            self.log("Top status transitions:")
            for label, count in sorted(transitions.items(), key=lambda kv: -kv[1])[:10]:
                self.log(f"  {label:<44}{count:>6,}")
        self.log("")
        self.log(f"Done - {result['written']:,} rows written.")
        self.status_text.set(
            f"Done - {result['written']:,} rows written, "
            f"{outcome_counts.get(OUT_OUTSTANDING, 0):,} still outstanding.")

        if self.open_when_done.get():
            try:
                if sys.platform.startswith("win"):
                    os.startfile(output)  # noqa: S606
                elif sys.platform == "darwin":
                    subprocess.Popen(["open", output])
                else:
                    subprocess.Popen(["xdg-open", output])
            except Exception as exc:
                self.log(f"Could not open the output file: {exc}")


def main():
    root = tk.Tk()
    root.title(APP_TITLE)
    root.geometry("1000x980")
    # Only the log stretches, so the floor is roughly the fixed content plus a
    # couple of visible log lines.
    root.minsize(900, 880)
    try:
        ttk.Style().theme_use("vista")
    except tk.TclError:
        pass
    PatchReportApp(root)
    if openpyxl is None:
        messagebox.showerror(APP_TITLE, "openpyxl is not installed.\n\nRun: pip install openpyxl")
    root.mainloop()


if __name__ == "__main__":
    main()
