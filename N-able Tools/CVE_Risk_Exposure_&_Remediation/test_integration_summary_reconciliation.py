"""
test_integration_summary_reconciliation.py — end-to-end integration test.

Builds a small, fully-known synthetic vulnerability export + RMM export
(+ a patch report), runs the real orchestrator pipeline against them,
opens the generated .xlsx, and checks that the Summary sheet's numbers
actually reconcile against what's written to the product sheets and the
health-score scope — instead of trusting that each layer (data_pipeline →
resolution → summary_sheet → product_sheets) agrees with the others just
because each one has its own unit tests.

Every prior bug fixed in this codebase (the row-misalignment bug, the
Top-At-Risk dedup-timing gap, the stale-only-product scoping gap, the
health-score scope gap) was a case where two sheets/tables that were
SUPPOSED to describe the same numbers quietly didn't. Unit tests on the
individual functions wouldn't have caught most of these — they only show
up when you actually build a workbook and compare what one sheet says
against what another sheet (or the raw input) says. That's what this file
does.

Dataset (see VULN_ROWS below):
  TestBrowser  — 3 rows, 2 devices, CVSS 9.2-9.5 → gets its own product
                 sheet at threshold 9.0 (1 resolved via raw status,
                 2 unresolved).
  TestOffice   — 2 rows, 1 device, CVSS 7.2-8.5 → below the 9.0 report
                 threshold, so NO product sheet — but IS inside the
                 health-score scope (floor = min(9.0, 7.0) = 7.0). This is
                 the exact scenario the health-score scope-gap bug lived in.
  Google Chrome — 1 row, resolved purely via PATCH EVIDENCE (blank raw
                 status, not RESOLVED) — proves patch-confirmed rows
                 reconcile through the workbook too, not just raw
                 Threat Status == RESOLVED. Deliberately given a blank
                 status rather than UNRESOLVED: orchestrator.py has a
                 "scanner override" that discards patch evidence whenever
                 the raw scanner explicitly says UNRESOLVED for that exact
                 pair (a real, separate safeguard against stale patch
                 records — see orchestrator.py's "Scanner override" log
                 line) — so a genuine patch-evidence test has to route
                 around that intentionally, not defeat it. Requires a
                 FIXED_VERSION_RULES entry to make the patch's version
                 register as compliant, monkeypatched for this test only.

Network calls (cve_lookup's live CVSS enrichment) are mocked out — this
test verifies the report-generation pipeline, not third-party CVE APIs.

Run with: pytest test_integration_summary_reconciliation.py -v
"""
import csv
import os
from datetime import datetime
from unittest.mock import patch

os.environ.setdefault('PYTEST_CURRENT_TEST', 'bootstrap')

import openpyxl
import pytest

import data_pipeline
from orchestrator import DashboardRequest, run as run_dashboard


# ── Synthetic dataset ──────────────────────────────────────────────────────────
# (Asset Name, Vulnerability ID, Affected Products, Threat Status, CVSS Score, Severity)
VULN_ROWS = [
    ('DEV001', 'CVE-2099-0001', 'TestBrowser',   'UNRESOLVED', 9.5, 'CRITICAL'),
    ('DEV001', 'CVE-2099-0002', 'TestBrowser',   'RESOLVED',   9.2, 'CRITICAL'),
    ('DEV002', 'CVE-2099-0001', 'TestBrowser',   'UNRESOLVED', 9.5, 'CRITICAL'),
    ('DEV001', 'CVE-2099-0003', 'TestOffice',    'UNRESOLVED', 8.5, 'IMPORTANT'),
    ('DEV001', 'CVE-2099-0004', 'TestOffice',    'RESOLVED',   7.2, 'IMPORTANT'),
    # Patch-evidence row — see module docstring for why the status is blank,
    # not UNRESOLVED. Fixed by a matching patch report row + a monkeypatched
    # FIXED_VERSION_RULES entry, not by the raw status column at all.
    ('DEV002', 'CVE-2099-0005', 'Google Chrome', '',           9.3, 'CRITICAL'),
]

PATCH_CVE_ID       = 'CVE-2099-0005'
PATCH_FIXED_VER    = '100.0'
PATCH_MATCHED_VER  = 'Google Chrome 110.0.5481.100'

# Expected outcomes, derived by hand from VULN_ROWS above — the point of
# this file is to check the generated workbook against these, not against
# whatever the pipeline happens to produce.
EXPECTED = {
    'testbrowser_resolved_rows':   1,   # CVE-2099-0002 on DEV001
    'testbrowser_unresolved_rows': 2,   # CVE-2099-0001 on DEV001 and DEV002
    'active_resolved_rows':    2,   # TestBrowser(1) + Google Chrome(1, patch evidence)
    'active_unresolved_rows':  2,   # TestBrowser(2) + Google Chrome(0)
    # Health scope (CVSS >= 7.0) spans all three products: 3 TestBrowser +
    # 2 TestOffice + 1 Google Chrome = 6 rows, 3 resolved
    # (CVE-0002, CVE-0004, CVE-0005) = 0.5.
    'health_scope_total_rows':    6,
    'health_scope_resolved_rows': 3,
    'health_resolution_rate':     0.5,
}


def _write_vuln_csv(path):
    now_date = datetime.now().strftime('%Y-%m-%d')
    with open(path, 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f)
        w.writerow(['Asset Name', 'Vulnerability ID', 'Affected Products', 'Threat Status',
                    'CVSS Score', 'Severity', 'Customer', 'Site', 'Has Exploit', 'CISA KEV',
                    'Last scanned', 'First detected', 'Date Published'])
        for name, cve, product, status, score, sev in VULN_ROWS:
            w.writerow([name, cve, product, status, score, sev,
                        'TestCo', 'HQ', 'No', 'No', now_date,
                        '2026-01-01', '2026-01-01'])


def _write_rmm_csv(path):
    now_dt = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with open(path, 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f)
        w.writerow(['Device Name', 'Last Response', 'OS', 'Device Type', 'Username'])
        w.writerow(['DEV001', now_dt, 'WINDOWS', 'Workstation', 'testuser1'])
        w.writerow(['DEV002', now_dt, 'WINDOWS', 'Workstation', 'testuser2'])


def _write_patch_csv(path):
    """One patch record confirming CVE-2099-0005 fixed on DEV002 — installed
    well after the CVE's (fake) publish date, with a version that satisfies
    the monkeypatched FIXED_VERSION_RULES entry below."""
    with open(path, 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f)
        w.writerow(['Client', 'Site', 'Device', 'Status', 'Patch', 'Discovered / Install Date'])
        w.writerow(['TestCo', 'HQ', 'DEV002', 'Installed', PATCH_MATCHED_VER, '2026-06-01'])


@pytest.fixture(scope='module')
def workbook_path(tmp_path_factory):
    """Build the synthetic inputs, run the real pipeline, return the output path."""
    tmp_dir    = tmp_path_factory.mktemp('integration')
    vuln_path  = tmp_dir / 'vuln.csv'
    rmm_path   = tmp_dir / 'rmm.csv'
    patch_path = tmp_dir / 'patch.csv'
    out_path   = tmp_dir / 'output.xlsx'

    _write_vuln_csv(vuln_path)
    _write_rmm_csv(rmm_path)
    _write_patch_csv(patch_path)

    req = DashboardRequest(
        vuln_path=str(vuln_path),
        output_path=str(out_path),
        rmm_path=str(rmm_path),
        skip_rmm=False,
        threshold=9.0,
        cutoff_date=None,
        show_all_dates=True,
        prev_report_path=None,
        include_trend=False,
        include_health_score=True,
        patch_path=str(patch_path),
        include_patch=True,
    )
    # Mock the live CVE-score enrichment network calls — this test verifies
    # report generation, not third-party API availability. Also inject a
    # fixed-version rule for the patch-evidence CVE so its installed patch
    # version registers as compliant, without touching the real config.json.
    with patch.dict(data_pipeline.FIXED_VERSION_RULES,
                    {'chrome': {PATCH_CVE_ID: PATCH_FIXED_VER}}, clear=False):
        with patch('cve_lookup.enrich_from_detections', return_value=0):
            result = run_dashboard(req)
    assert result.success, f"Dashboard generation failed: {result.message}"
    return str(out_path)


@pytest.fixture(scope='module')
def wb_values(workbook_path):
    """Workbook with cached formula results (for reading numbers)."""
    return openpyxl.load_workbook(workbook_path, data_only=True)


@pytest.fixture(scope='module')
def wb_formulas(workbook_path):
    """Workbook with raw formula strings (for checking live vs static cells)."""
    return openpyxl.load_workbook(workbook_path, data_only=False)


def _sheet_rows(ws):
    """
    Return a list of {header: value} dicts for every real data row in a
    product sheet — excludes the legend section at the bottom, which also
    has non-empty first-column values ('Legend', '  (blue row)', etc.) and
    would otherwise be miscounted as data rows.

    Handles both product sheet layouts: a normal triage sheet (header on
    row 1) and the "Patch Confirmed" variant used when every row on that
    sheet is already ☑ (banner + note before the header, so the header
    isn't necessarily on row 1). Search the first few rows for whichever
    one actually has 'Vulnerability Name' in it, rather than assuming row 1.
    """
    header = None
    header_row_idx = None
    for hr in range(1, 5):
        vals = [c.value for c in ws[hr]]
        if vals and 'Vulnerability Name' in vals and 'Name' in vals:
            header = vals
            header_row_idx = hr
            break
    assert header is not None, f"Could not find a header row in sheet '{ws.title}'"

    vn_idx = header.index('Vulnerability Name')
    rows = []
    for r in range(header_row_idx + 1, ws.max_row + 1):
        if ws.cell(row=r, column=vn_idx + 1).value is None:
            continue
        rows.append({header[i]: ws.cell(row=r, column=i + 1).value for i in range(len(header))})
    return rows


def _summary_cell(ws, row_label, col):
    """Find a row by its first-column label and return a given column's value."""
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == row_label:
            return ws.cell(row=r, column=col).value
    return None


# ── Product sheet scope ───────────────────────────────────────────────────────

class TestProductSheetScope:
    def test_testbrowser_sheet_exists(self, wb_values):
        assert 'TestBrowser' in wb_values.sheetnames

    def test_google_chrome_sheet_exists(self, wb_values):
        """CVSS 9.3 clears the report threshold, so this gets its own sheet
        even though it has no raw status contributing to its ☑/☐ verdict."""
        assert 'Google Chrome' in wb_values.sheetnames

    def test_testoffice_has_no_dedicated_sheet(self, wb_values):
        """TestOffice has no rows at the 9.0 report threshold — it must not
        get a product sheet, even though it's inside the health scope."""
        assert 'TestOffice' not in wb_values.sheetnames

    def test_testbrowser_row_count_and_resolution(self, wb_values):
        rows = _sheet_rows(wb_values['TestBrowser'])
        resolved   = [r for r in rows if r.get('Resolved') == '☑']
        unresolved = [r for r in rows if r.get('Resolved') == '☐']
        assert len(resolved)   == EXPECTED['testbrowser_resolved_rows']
        assert len(unresolved) == EXPECTED['testbrowser_unresolved_rows']

    def test_patch_evidence_resolves_row_with_blank_raw_status(self, wb_values):
        """
        CVE-2099-0005 has a blank raw status — nothing from source 2 (raw
        status) contributes to its resolution — so a ☑ here can only have
        come from patch evidence (source 1). Proves patch-confirmed rows
        reconcile through to the actual product sheet, not just rows whose
        raw Threat Status literally says RESOLVED.
        """
        rows = _sheet_rows(wb_values['Google Chrome'])
        matching = [r for r in rows if r.get('Vulnerability Name') == PATCH_CVE_ID]
        assert len(matching) == 1, f"Expected exactly one {PATCH_CVE_ID} row"
        assert matching[0]['Resolved'] == '☑', (
            f"{PATCH_CVE_ID} has blank raw status — its ☑ must come from patch "
            f"evidence alone, but it shows {matching[0]['Resolved']!r}"
        )


# ── Summary reconciliation against product sheets ─────────────────────────────

class TestSummaryReconciliation:
    def test_resolution_status_matches_product_sheets(self, wb_values):
        """
        Resolution Status (Active) must equal the actual ☑/☐ split summed
        across BOTH active product sheets (TestBrowser + Google Chrome) —
        including the patch-evidence-resolved row, which has no raw status
        of its own to be counted by a naive "check Threat Status" approach.
        """
        ws = wb_values['Summary']
        all_rows = _sheet_rows(wb_values['TestBrowser']) + _sheet_rows(wb_values['Google Chrome'])
        actual_resolved   = sum(1 for r in all_rows if r.get('Resolved') == '☑')
        actual_unresolved = sum(1 for r in all_rows if r.get('Resolved') == '☐')
        assert _summary_cell(ws, 'Resolved',   2) == actual_resolved   == EXPECTED['active_resolved_rows']
        assert _summary_cell(ws, 'Unresolved', 2) == actual_unresolved == EXPECTED['active_unresolved_rows']

    def test_top_at_risk_matches_product_sheet_unresolved_count(self, wb_values):
        """
        Each device's 'Unresolved CVEs' count in Top At-Risk Devices must
        equal its actual ☐ count across all active product sheets. This is
        exactly the class of bug found earlier: Top At-Risk used to compute
        this independently from raw Threat Status, and could disagree with
        what the product sheets actually showed.
        """
        wsum = wb_values['Summary']
        top_at_risk = {}
        header_row = None
        for r in range(1, wsum.max_row + 1):
            if wsum.cell(row=r, column=1).value == '💻 Device Name':
                header_row = r
                break
        assert header_row is not None, "Top At-Risk Devices table not found"
        r = header_row + 1
        while wsum.cell(row=r, column=1).value and not str(wsum.cell(row=r, column=1).value).startswith('ℹ'):
            top_at_risk[wsum.cell(row=r, column=1).value] = wsum.cell(row=r, column=3).value
            r += 1

        all_rows = _sheet_rows(wb_values['TestBrowser']) + _sheet_rows(wb_values['Google Chrome'])
        for device in ('DEV001', 'DEV002'):
            actual_unresolved_cves = {
                row['Vulnerability Name'] for row in all_rows
                if row.get('Name') == device and row.get('Resolved') == '☐'
            }
            if not actual_unresolved_cves:
                assert device not in top_at_risk, (
                    f"{device} has zero unresolved CVEs but still appears in Top At-Risk Devices"
                )
                continue
            assert device in top_at_risk, f"{device} missing from Top At-Risk Devices"
            assert top_at_risk[device] == len(actual_unresolved_cves), (
                f"{device}: Top At-Risk shows {top_at_risk[device]}, "
                f"product sheets show {len(actual_unresolved_cves)}"
            )

    def test_key_metrics_total_rows_matches_product_sheet_row_count(self, wb_values):
        """'Total detection rows' (Active) must equal the actual row count
        summed across both active product sheets."""
        ws = wb_values['Summary']
        total_rows = len(_sheet_rows(wb_values['TestBrowser'])) + len(_sheet_rows(wb_values['Google Chrome']))
        assert _summary_cell(ws, 'Total detection rows', 3) == total_rows


# ── Health-score scope gap ─────────────────────────────────────────────────────

class TestHealthScoreScope:
    def test_health_score_section_present(self, wb_values):
        ws = wb_values['Summary']
        assert any(
            ws.cell(row=r, column=1).value and 'Patching Health Score' in str(ws.cell(row=r, column=1).value)
            for r in range(1, ws.max_row + 1)
        )

    def test_resolution_rate_includes_products_with_no_sheet(self, wb_values):
        """
        The core regression check: Resolution rate's coverage fraction must
        reflect ALL THREE products (6 rows, 3 resolved = 0.5) — not just the
        products with their own sheet, which is what it would show if
        TestOffice's rows were silently dropped for having no product sheet.
        This is the exact health-score scope-gap bug.
        """
        ws = wb_values['Summary']
        coverage = _summary_cell(ws, 'Resolution rate', 2)
        assert coverage == pytest.approx(EXPECTED['health_resolution_rate'], abs=1e-6), (
            f"Resolution rate = {coverage}; expected "
            f"{EXPECTED['health_scope_resolved_rows']}/{EXPECTED['health_scope_total_rows']} "
            f"= {EXPECTED['health_resolution_rate']}. A lower value here means "
            f"TestOffice (no product sheet) was silently excluded from the health scope."
        )

    def test_health_score_is_static_when_hidden_products_exist(self, wb_formulas):
        """
        TestOffice has no dedicated sheet, so a live COUNTIF formula has
        nothing to reference for it. The final score cell must be a plain
        static value, not a formula — otherwise the score would silently
        drop TestOffice the moment Excel recalculates it.
        """
        ws = wb_formulas['Summary']
        score_row = None
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=1).value
            if v and str(v).startswith('Patching Health Score  (Grade'):
                score_row = r
                break
        assert score_row is not None, "Final Health Score row not found"
        cell = ws.cell(row=score_row, column=3)
        assert not (isinstance(cell.value, str) and cell.value.startswith('=')), (
            "Health Score cell is a live formula, but TestOffice has no product "
            "sheet for a COUNTIF to reference — it should be static."
        )
        assert isinstance(cell.value, (int, float))

    def test_footnote_does_not_claim_live_when_score_is_static(self, wb_values):
        """
        The footnote's "⚡ Blue cells update automatically..." claim must be
        absent whenever the score is static (see previous test) — otherwise
        a reader could toggle a checkbox expecting the score to follow,
        when it silently won't for the hidden-product component.
        """
        ws = wb_values['Summary']
        footnote_text = ''
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=1).value
            if v and 'Grade bands:' in str(v):
                footnote_text = str(v)
                break
        assert footnote_text, "Health Score footnote row not found"
        assert 'update automatically' not in footnote_text.lower(), (
            "Footnote claims cells update automatically, but the score is static "
            "(TestOffice has no product sheet for a live formula to reference)."
        )