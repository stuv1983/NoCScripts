"""
RMM Device Inventory Duplicate Cleanup  —  Tkinter GUI
=======================================================
Compares N-Sight device inventories for two RMM groups.
Identifies devices with matching serial numbers, confirms true
duplicates, and documents removed Pop-Up Health records.

Run:  python rmm_duplicate_cleanup.py
"""

import sys
import threading
import subprocess
from datetime import datetime, timezone
from pathlib import Path
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Core config ─────────────────────────────────────────────────────────────
SERIAL_COL   = "Serial number"
HOSTNAME_COL = "Device name"
CONFIRM_COLS = ["Manufacturer", "Model", "Processor"]
LOG_COLS = [
    "Serial number", "Device name", "Customer name", "Site name",
    "Device type", "Manufacturer", "Model", "MAC Address",
    "Last response (UTC)", "Username",
]

# ── Excel styles ─────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", start_color="1F497D", end_color="1F497D")
REMOVED_FILL  = PatternFill("solid", start_color="FCE4D6", end_color="FCE4D6")
REJECTED_FILL = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
SUMMARY_FILL  = PatternFill("solid", start_color="DEEAF1", end_color="DEEAF1")
THIN   = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ── Core logic (unchanged from CLI version) ──────────────────────────────────

def load_inventory(path: str) -> pd.DataFrame:
    df = pd.read_excel(path, dtype=str)
    df.columns = df.columns.str.strip()
    df[SERIAL_COL] = df[SERIAL_COL].fillna("").str.strip().str.upper()
    return df


def find_serial_matches(ref_df: pd.DataFrame, target_df: pd.DataFrame) -> pd.DataFrame:
    ref_serials = set(ref_df[SERIAL_COL][ref_df[SERIAL_COL] != ""])
    return target_df[target_df[SERIAL_COL].isin(ref_serials)].copy()


def confirm_duplicates(candidates, ref_df):
    ref_idx = ref_df.set_index(SERIAL_COL)
    confirmed, rejected = [], []
    for _, row in candidates.iterrows():
        serial = row[SERIAL_COL]
        if serial not in ref_idx.index:
            rejected.append(row)
            continue
        ref_match = ref_idx.loc[serial]
        if isinstance(ref_match, pd.DataFrame):
            ref_match = ref_match.iloc[0]
        mismatches = []
        for col in CONFIRM_COLS:
            if col not in row.index or col not in ref_match.index:
                continue
            if (str(row[col]).strip().upper() != str(ref_match[col]).strip().upper()
                    and str(row[col]).strip() and str(ref_match[col]).strip()):
                mismatches.append(f"{col}: target='{row[col]}' vs ref='{ref_match[col]}'")
        r = row.copy()
        if mismatches:
            r["_rejection_reason"] = "; ".join(mismatches)
            rejected.append(r)
        else:
            r["_ref_hostname"] = ref_match.get(HOSTNAME_COL, "")
            confirmed.append(r)
    confirmed_df = pd.DataFrame(confirmed) if confirmed else pd.DataFrame(columns=candidates.columns)
    rejected_df  = pd.DataFrame(rejected)  if rejected  else pd.DataFrame(columns=candidates.columns)
    return confirmed_df, rejected_df


def remove_duplicates(target_df, confirmed_df):
    return target_df[~target_df[SERIAL_COL].isin(set(confirmed_df[SERIAL_COL]))].copy()


def _style_header(ws, row_num, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = BORDER


def _autofit(ws, min_w=12, max_w=50):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width  = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[letter].width = min(max(width + 2, min_w), max_w)


def _write_df(ws, df, fill=None, start_row=1):
    for ci, name in enumerate(df.columns, 1):
        ws.cell(row=start_row, column=ci, value=name)
    _style_header(ws, start_row, len(df.columns))
    for ri, (_, row) in enumerate(df.iterrows(), start_row + 1):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical="top")
            cell.border    = BORDER
            if fill:
                cell.fill = fill


def write_report(output_path, confirmed, rejected, cleaned_target,
                 ref_df, raw_target, ref_label, target_label):
    ts = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    wb = openpyxl.Workbook()

    # Summary
    ws = wb.active
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 28
    title_cell = ws.cell(1, 1, "RMM Duplicate Cleanup Report")
    title_cell.font = Font(bold=True, size=14, name="Arial")
    title_cell.fill = SUMMARY_FILL
    rows = [
        ("Report generated (UTC)",            ts),
        (f"{ref_label} total devices",         len(ref_df)),
        (f"{target_label} total devices",      len(raw_target)),
        ("Serial number matches found",        len(confirmed) + len(rejected)),
        ("Confirmed true duplicates",          len(confirmed)),
        ("Rejected (hardware mismatch)",       len(rejected)),
        (f"{target_label} devices removed",    len(confirmed)),
        (f"{target_label} devices remaining",  len(cleaned_target)),
    ]
    for r, (label, val) in enumerate(rows, 3):
        a = ws.cell(r, 1, label)
        b = ws.cell(r, 2, val)
        for cell in (a, b):
            cell.font   = Font(name="Arial", size=10)
            cell.border = BORDER
        a.font = Font(name="Arial", size=10, bold=True)
        b.fill = SUMMARY_FILL

    # Removed Devices
    ws_rem = wb.create_sheet("Removed Devices")
    if len(confirmed):
        cols   = [c for c in LOG_COLS if c in confirmed.columns] + \
                 (["_ref_hostname"] if "_ref_hostname" in confirmed.columns else [])
        df_out = confirmed[cols].copy()
        df_out.rename(columns={"_ref_hostname": f"Matched {ref_label} hostname"}, inplace=True)
        _write_df(ws_rem, df_out, fill=REMOVED_FILL)
        _autofit(ws_rem)
    else:
        ws_rem.cell(1, 1, "No confirmed duplicates found.").font = Font(name="Arial", italic=True)

    # Rejected Candidates
    ws_rej = wb.create_sheet("Rejected Candidates")
    if len(rejected):
        cols   = [c for c in LOG_COLS if c in rejected.columns] + \
                 (["_rejection_reason"] if "_rejection_reason" in rejected.columns else [])
        df_out = rejected[cols].copy()
        df_out.rename(columns={"_rejection_reason": "Rejection reason"}, inplace=True)
        _write_df(ws_rej, df_out, fill=REJECTED_FILL)
        _autofit(ws_rej)
    else:
        ws_rej.cell(1, 1, "No rejected candidates.").font = Font(name="Arial", italic=True)

    # Cleaned target inventory
    ws_clean = wb.create_sheet(f"{target_label} (Cleaned)")
    if len(cleaned_target):
        _write_df(ws_clean, cleaned_target.reset_index(drop=True))
        _autofit(ws_clean)

    wb.save(output_path)
    return output_path


# ── GUI ───────────────────────────────────────────────────────────────────────

ACCENT   = "#1F497D"
BG       = "#F4F6F9"
CARD_BG  = "#FFFFFF"
TEXT     = "#1A1A2E"
MUTED    = "#6B7280"
SUCCESS  = "#16A34A"
WARNING  = "#D97706"
DANGER   = "#DC2626"
RADIUS   = 6


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("RMM Duplicate Cleanup")
        self.resizable(False, False)
        self.configure(bg=BG)

        # State
        self.ref_path    = tk.StringVar()
        self.target_path = tk.StringVar()
        self.out_path    = tk.StringVar()
        self.ref_label   = tk.StringVar(value="Lumia Care")
        self.target_label= tk.StringVar(value="Pop-Up Health")
        self._results    = None

        self._build_ui()
        self._centre()

    # ── Layout ───────────────────────────────────────────────────────────────

    def _build_ui(self):
        # ── Header ──
        hdr = tk.Frame(self, bg=ACCENT, pady=14)
        hdr.pack(fill="x")
        tk.Label(hdr, text="RMM Device Inventory", bg=ACCENT, fg="white",
                 font=("Segoe UI", 16, "bold")).pack()
        tk.Label(hdr, text="Duplicate Cleanup Tool", bg=ACCENT, fg="#BDD7EE",
                 font=("Segoe UI", 10)).pack()

        body = tk.Frame(self, bg=BG, padx=24, pady=20)
        body.pack(fill="both", expand=True)

        # ── File pickers ──
        card = self._card(body, "Input Files")
        card.pack(fill="x", pady=(0, 12))

        self._file_row(card, "Reference group (Lumia Care):", self.ref_label,
                       self.ref_path, 0)
        self._file_row(card, "Target group (Pop-Up Health):", self.target_label,
                       self.target_path, 1)

        # ── Output ──
        out_card = self._card(body, "Output Report")
        out_card.pack(fill="x", pady=(0, 12))
        self._output_row(out_card)

        # ── Run button ──
        btn_frame = tk.Frame(body, bg=BG)
        btn_frame.pack(fill="x", pady=(4, 0))

        self.run_btn = tk.Button(
            btn_frame, text="▶  Run Duplicate Check",
            command=self._run,
            bg=ACCENT, fg="white", activebackground="#163A6B",
            font=("Segoe UI", 11, "bold"),
            relief="flat", cursor="hand2",
            padx=20, pady=10,
        )
        self.run_btn.pack(side="left")

        self.open_btn = tk.Button(
            btn_frame, text="📂  Open Report",
            command=self._open_report,
            bg="#E2EAF4", fg=ACCENT, activebackground="#C7D7EC",
            font=("Segoe UI", 10),
            relief="flat", cursor="hand2",
            padx=14, pady=10,
            state="disabled",
        )
        self.open_btn.pack(side="left", padx=(10, 0))

        # ── Log / progress ──
        log_card = self._card(body, "Log")
        log_card.pack(fill="both", expand=True, pady=(12, 0))

        self.log = tk.Text(log_card, height=14, bg="#0D1117", fg="#E6EDF3",
                           font=("Consolas", 9), relief="flat",
                           state="disabled", wrap="word")
        self.log.pack(fill="both", expand=True, padx=2, pady=2)
        sb = ttk.Scrollbar(self.log, command=self.log.yview)
        self.log["yscrollcommand"] = sb.set

        # colour tags
        self.log.tag_config("ok",   foreground="#3FB950")
        self.log.tag_config("warn", foreground="#D29922")
        self.log.tag_config("err",  foreground="#F85149")
        self.log.tag_config("info", foreground="#79C0FF")
        self.log.tag_config("dim",  foreground="#8B949E")

        # ── Status bar ──
        self.status_var = tk.StringVar(value="Ready")
        tk.Label(self, textvariable=self.status_var, bg=BG, fg=MUTED,
                 font=("Segoe UI", 8), anchor="w", padx=24
                 ).pack(fill="x", pady=(4, 6))

    def _card(self, parent, title):
        outer = tk.Frame(parent, bg=CARD_BG, relief="flat",
                         highlightbackground="#D1D5DB", highlightthickness=1)
        tk.Label(outer, text=title, bg=CARD_BG, fg=ACCENT,
                 font=("Segoe UI", 9, "bold"), anchor="w",
                 padx=12, pady=6).pack(fill="x")
        ttk.Separator(outer, orient="horizontal").pack(fill="x")
        return outer

    def _file_row(self, parent, prompt, label_var, path_var, grid_row):
        row = tk.Frame(parent, bg=CARD_BG, pady=4)
        row.pack(fill="x", padx=12, pady=2)

        tk.Label(row, text=prompt, bg=CARD_BG, fg=TEXT,
                 font=("Segoe UI", 9), width=34, anchor="w").grid(
                     row=0, column=0, sticky="w")

        name_entry = tk.Entry(row, textvariable=label_var, width=18,
                              font=("Segoe UI", 9), relief="solid", bd=1)
        name_entry.grid(row=0, column=1, padx=(0, 8), sticky="w")

        path_entry = tk.Entry(row, textvariable=path_var, width=36,
                              font=("Segoe UI", 9), relief="solid", bd=1,
                              state="readonly")
        path_entry.grid(row=0, column=2, padx=(0, 6), sticky="ew")

        tk.Button(row, text="Browse…",
                  command=lambda pv=path_var: self._browse_xlsx(pv),
                  bg="#E2EAF4", fg=ACCENT, relief="flat", cursor="hand2",
                  font=("Segoe UI", 9), padx=8
                  ).grid(row=0, column=3)

        row.columnconfigure(2, weight=1)

    def _output_row(self, parent):
        row = tk.Frame(parent, bg=CARD_BG, pady=4)
        row.pack(fill="x", padx=12, pady=2)

        tk.Label(row, text="Save report to:", bg=CARD_BG, fg=TEXT,
                 font=("Segoe UI", 9), width=34, anchor="w").grid(
                     row=0, column=0, sticky="w")

        # Set a sensible default
        default_out = str(Path.home() / "Desktop" / "duplicate_removal_report.xlsx")
        self.out_path.set(default_out)

        path_entry = tk.Entry(row, textvariable=self.out_path, width=54,
                              font=("Segoe UI", 9), relief="solid", bd=1,
                              state="readonly")
        path_entry.grid(row=0, column=1, padx=(0, 6), sticky="ew")

        tk.Button(row, text="Browse…",
                  command=self._browse_save,
                  bg="#E2EAF4", fg=ACCENT, relief="flat", cursor="hand2",
                  font=("Segoe UI", 9), padx=8
                  ).grid(row=0, column=2)

        row.columnconfigure(1, weight=1)

    # ── Dialogs ──────────────────────────────────────────────────────────────

    def _browse_xlsx(self, path_var):
        p = filedialog.askopenfilename(
            title="Select inventory report",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
        )
        if p:
            path_var.set(p)

    def _browse_save(self):
        p = filedialog.asksaveasfilename(
            title="Save report as",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="duplicate_removal_report.xlsx",
        )
        if p:
            self.out_path.set(p)

    # ── Logging ──────────────────────────────────────────────────────────────

    def _log(self, msg, tag=""):
        self.log.configure(state="normal")
        self.log.insert("end", msg + "\n", tag)
        self.log.see("end")
        self.log.configure(state="disabled")

    def _clear_log(self):
        self.log.configure(state="normal")
        self.log.delete("1.0", "end")
        self.log.configure(state="disabled")

    # ── Run ──────────────────────────────────────────────────────────────────

    def _run(self):
        ref_path    = self.ref_path.get().strip()
        target_path = self.target_path.get().strip()
        out_path    = self.out_path.get().strip()
        ref_label   = self.ref_label.get().strip() or "Reference"
        target_label= self.target_label.get().strip() or "Target"

        if not ref_path:
            messagebox.showwarning("Missing file", "Please select the Reference group file.")
            return
        if not target_path:
            messagebox.showwarning("Missing file", "Please select the Target group file.")
            return
        if not out_path:
            messagebox.showwarning("Missing path", "Please choose an output path.")
            return

        self.run_btn.configure(state="disabled", text="Running…")
        self.open_btn.configure(state="disabled")
        self._clear_log()
        self.status_var.set("Working…")
        self._results = None

        thread = threading.Thread(
            target=self._worker,
            args=(ref_path, target_path, out_path, ref_label, target_label),
            daemon=True,
        )
        thread.start()

    def _worker(self, ref_path, target_path, out_path, ref_label, target_label):
        try:
            self._log(f"{'='*54}", "dim")
            self._log(f"  RMM Duplicate Cleanup", "info")
            self._log(f"{'='*54}", "dim")

            self._log(f"\nLoading {ref_label}…", "dim")
            ref_df = load_inventory(ref_path)
            self._log(f"  ✔ {len(ref_df):,} devices loaded", "ok")

            self._log(f"Loading {target_label}…", "dim")
            target_raw = load_inventory(target_path)
            self._log(f"  ✔ {len(target_raw):,} devices loaded", "ok")

            self._log("\nSearching for serial number matches…", "dim")
            candidates = find_serial_matches(ref_df, target_raw)
            self._log(f"  → {len(candidates)} candidate(s) found", "info")

            if candidates.empty:
                self._log("\n✔ No duplicates found — nothing to remove.", "ok")
                self.after(0, self._done, None, out_path)
                return

            self._log("Confirming true duplicates…", "dim")
            confirmed, rejected = confirm_duplicates(candidates, ref_df)

            self._log(f"  ✔ {len(confirmed)} confirmed duplicate(s)", "ok")
            if len(rejected):
                self._log(f"  ⚠ {len(rejected)} rejected (hardware mismatch — NOT removed)", "warn")

            cleaned = remove_duplicates(target_raw, confirmed)

            if not confirmed.empty:
                self._log(f"\n{'─'*54}", "dim")
                self._log(f"  DEVICES TO BE REMOVED FROM {target_label.upper()}", "warn")
                self._log(f"{'─'*54}", "dim")
                for _, row in confirmed.iterrows():
                    self._log(
                        f"  Hostname : {row.get(HOSTNAME_COL, 'N/A')}\n"
                        f"  Serial   : {row.get(SERIAL_COL, 'N/A')}\n"
                        f"  Model    : {row.get('Manufacturer','')} {row.get('Model','')}\n"
                        f"  Matches  : {row.get('_ref_hostname','N/A')} in {ref_label}\n",
                        "warn"
                    )

            self._log("Writing Excel report…", "dim")
            write_report(out_path, confirmed, rejected, cleaned,
                         ref_df, target_raw, ref_label, target_label)
            self._log(f"\n✔ Report saved to:\n  {out_path}", "ok")

            self.after(0, self._done, (confirmed, rejected, cleaned), out_path)

        except Exception as exc:
            self._log(f"\n✖ ERROR: {exc}", "err")
            self.after(0, self._error)

    def _done(self, results, out_path):
        self._results = results
        self.run_btn.configure(state="normal", text="▶  Run Duplicate Check")
        if results is not None:
            confirmed, rejected, cleaned = results
            self.open_btn.configure(state="normal")
            self.status_var.set(
                f"Done — {len(confirmed)} removed, {len(rejected)} rejected, "
                f"{len(cleaned)} devices remaining in target group."
            )
        else:
            self.status_var.set("Done — no duplicates found.")

    def _error(self):
        self.run_btn.configure(state="normal", text="▶  Run Duplicate Check")
        self.status_var.set("Error — see log above.")

    def _open_report(self):
        path = self.out_path.get()
        if not path or not Path(path).exists():
            messagebox.showerror("File not found", "Report file does not exist.")
            return
        if sys.platform == "win32":
            import os; os.startfile(path)
        elif sys.platform == "darwin":
            subprocess.call(["open", path])
        else:
            subprocess.call(["xdg-open", path])

    # ── Helpers ──────────────────────────────────────────────────────────────

    def _centre(self):
        self.update_idletasks()
        w, h = self.winfo_width(), self.winfo_height()
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        self.geometry(f"+{(sw-w)//2}+{(sh-h)//2}")


if __name__ == "__main__":
    app = App()
    app.mainloop()